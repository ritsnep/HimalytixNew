import datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import Client, SimpleTestCase, TestCase
from django.urls import reverse

from accounting.models import Organization, ReportDefinition
from accounting.services.report_export_service import ReportExportService
from accounting.services.report_service import ReportService

User = get_user_model()


class ReportServiceStoredProcedureTests(TestCase):
    """Smoke tests for stored-procedure backed report service."""

    def setUp(self):
        self.organization = Organization.objects.create(name="Acme Corp", code="ACME")
        self.service = ReportService(self.organization)

    def test_general_ledger_returns_structure_without_data(self):
        """General ledger should return the expected keys even when no data exists."""
        start = datetime.date.today() - datetime.timedelta(days=7)
        end = datetime.date.today()
        self.service.set_date_range(start, end)
        report = self.service.generate_general_ledger()
        self.assertEqual(report["report_type"], "general_ledger")
        self.assertEqual(report["lines"], [])
        self.assertIn("totals", report)
        self.assertEqual(report["totals"]["total_debit"], Decimal("0.00"))

    def test_trial_balance_returns_structure_without_data(self):
        """Trial balance should execute stored function without seeded data."""
        as_of = datetime.date.today()
        report = self.service.generate_trial_balance(as_of)
        self.assertEqual(report["report_type"], "trial_balance")
        self.assertEqual(report["lines"], [])
        self.assertTrue(report["is_balanced"])

    def test_run_custom_definition(self):
        """Custom stored-procedure definitions should execute in order."""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                CREATE OR REPLACE FUNCTION fn_test_custom_report(p_input TEXT)
                RETURNS TABLE(result TEXT)
                LANGUAGE SQL
                AS $$
                    SELECT COALESCE(p_input, '') || '_ok' AS result;
                $$;
                """
            )

        definition = ReportDefinition.objects.create(
            organization=None,
            code="test_custom",
            name="Test Custom",
            stored_procedure="fn_test_custom_report",
            parameter_schema={"parameters": [{"name": "p_input", "type": "string", "required": True}]},
        )

        result = self.service.run_custom_definition(definition, {"p_input": "hello"})
        self.assertEqual(result["rows"], [{"result": "hello_ok"}])
        self.assertEqual(result["columns"], ["result"])

        with connection.cursor() as cursor:
            cursor.execute("DROP FUNCTION IF EXISTS fn_test_custom_report(TEXT);")


class ReportViewSmokeTests(TestCase):
    """Ensure report views render for authenticated users."""

    def setUp(self):
        self.organization = Organization.objects.create(name="Acme Corp", code="ACME")
        self.user = User.objects.create_user(
            username="reporter",
            email="reporter@example.com",
            password="secret123",
            organization=self.organization,
        )
        self.client = Client()
        self.client.login(username="reporter", password="secret123")

    def test_report_list_view(self):
        response = self.client.get(reverse("accounting:report_list"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "accounting/reports/report_list.html")

    def test_general_ledger_view(self):
        response = self.client.get(reverse("accounting:report_ledger"))
        self.assertEqual(response.status_code, 200)

    def test_trial_balance_view(self):
        response = self.client.get(reverse("accounting:report_trial_balance"))
        self.assertEqual(response.status_code, 200)

    def test_profit_loss_view(self):
        response = self.client.get(reverse("accounting:report_pl"))
        self.assertEqual(response.status_code, 200)

    def test_balance_sheet_view(self):
        response = self.client.get(reverse("accounting:report_bs"))
        self.assertEqual(response.status_code, 200)

    def test_cash_flow_view(self):
        response = self.client.get(reverse("accounting:report_cf"))
        self.assertEqual(response.status_code, 200)

    def test_ar_aging_view(self):
        response = self.client.get(reverse("accounting:report_ar_aging"))
        self.assertEqual(response.status_code, 200)

    def test_custom_report_view(self):
        definition = ReportDefinition.objects.create(
            organization=None,
            code="empty_custom",
            name="Empty Custom",
            stored_procedure="fn_report_trial_balance",
            parameter_schema={"parameters": [{"name": "p_org_id", "type": "int", "required": True}, {"name": "p_as_of", "type": "date", "required": True}]},
        )
        url = reverse("accounting:custom_report", args=[definition.code])
        response = self.client.get(url, {"p_org_id": self.organization.id, "p_as_of": datetime.date.today().isoformat()})
        self.assertEqual(response.status_code, 200)


class ReportExportViewTests(TestCase):
    """Ensure export endpoint validates parameters and handles format errors."""

    def setUp(self):
        self.organization = Organization.objects.create(name="Acme Corp", code="ACME")
        self.user = User.objects.create_user(
            username="exporter",
            email="exporter@example.com",
            password="secret123",
            organization=self.organization,
        )
        self.client = Client()
        self.client.login(username="exporter", password="secret123")

    def test_trial_balance_csv_export(self):
        response = self.client.post(
            reverse("accounting:report_export"),
            {
                "report_type": "trial_balance",
                "export_format": "csv",
                "as_of_date": datetime.date.today().isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])

    def test_invalid_report_type(self):
        response = self.client.post(
            reverse("accounting:report_export"),
            {"report_type": "unknown", "export_format": "csv"},
        )
        self.assertEqual(response.status_code, 400)


class ReportExportServiceTests(SimpleTestCase):
    """Validate export service formatting for stored-procedure report payloads."""

    def test_profit_and_loss_csv_contains_sections_and_totals(self):
        report_data = {
            "report_type": "profit_loss",
            "organization": "Acme Corp",
            "period": "2024-01-01 - 2024-01-31",
            "generated_at": datetime.datetime.now(),
            "sections": [
                {
                    "category": "Revenue",
                    "accounts": [
                        {"account_code": "4000", "account_name": "Sales", "debit": Decimal("0.00"), "credit": Decimal("5000.00"), "net": Decimal("5000.00")},
                    ],
                    "total": Decimal("5000.00"),
                },
                {
                    "category": "Expenses",
                    "accounts": [
                        {"account_code": "5000", "account_name": "Rent", "debit": Decimal("2000.00"), "credit": Decimal("0.00"), "net": Decimal("2000.00")},
                    ],
                    "total": Decimal("2000.00"),
                },
            ],
            "totals": {
                "total_income": Decimal("5000.00"),
                "total_expense": Decimal("2000.00"),
                "net_profit": Decimal("3000.00"),
            },
        }

        buffer, filename = ReportExportService.to_csv(report_data)
        content = buffer.getvalue().decode("utf-8")
        self.assertIn("Revenue", content)
        self.assertIn("Subtotal", content)
        self.assertIn("Net Profit", content)
        self.assertTrue(filename.endswith(".csv"))

    def test_balance_sheet_csv_groups_accounts(self):
        report_data = {
            "report_type": "balance_sheet",
            "organization": "Acme Corp",
            "as_of_date": datetime.date(2024, 1, 31),
            "generated_at": datetime.datetime.now(),
            "lines": [
                {
                    "account_code": "1000",
                    "account_name": "Cash",
                    "line_type": "asset",
                    "category": "Current Assets",
                    "amount": Decimal("1000.00"),
                },
                {
                    "account_code": "2000",
                    "account_name": "Accounts Payable",
                    "line_type": "liability",
                    "category": "Current Liabilities",
                    "amount": Decimal("250.00"),
                },
            ],
            "totals": {
                "total_assets": Decimal("1000.00"),
                "total_liabilities": Decimal("250.00"),
                "total_equity": Decimal("750.00"),
                "total_liabilities_equity": Decimal("1000.00"),
                "difference": Decimal("0.00"),
            },
        }

        buffer, _ = ReportExportService.to_csv(report_data)
        content = buffer.getvalue().decode("utf-8")
        self.assertIn("Current Assets", content)
        self.assertIn("Subtotal", content)
        self.assertIn("Total Assets", content)

    def test_custom_report_generic_export(self):
        report_data = {
            "report_type": "custom:sales_summary",
            "name": "Sales Summary",
            "organization": "Acme Corp",
            "generated_at": datetime.datetime.now(),
            "columns": ["Region", "Total"],
            "rows": [{"Region": "North", "Total": 5}, {"Region": "South", "Total": 7}],
        }

        csv_buffer, csv_filename = ReportExportService.to_csv(report_data)
        csv_content = csv_buffer.getvalue().decode("utf-8")
        self.assertIn("Region,Total", csv_content)
        self.assertIn("North,5", csv_content)
        self.assertTrue(csv_filename.startswith("custom:sales_summary"))

        excel_buffer, _ = ReportExportService.to_excel(report_data)
        excel_buffer.seek(0)
        from openpyxl import load_workbook  # Imported lazily for tests

        workbook = load_workbook(excel_buffer)
        sheet = workbook.active
        rows = [[cell.value for cell in row[:2]] for row in sheet.iter_rows(min_row=1, max_col=2)]
        self.assertIn(["Region", "Total"], rows)
        self.assertIn(["North", 5], rows)

    def test_invalid_export_format(self):
        response = self.client.post(
            reverse("accounting:report_export"),
            {
                "report_type": "trial_balance",
                "export_format": "invalid",
                "as_of_date": datetime.date.today().isoformat(),
            },
        )
        self.assertEqual(response.status_code, 400)
