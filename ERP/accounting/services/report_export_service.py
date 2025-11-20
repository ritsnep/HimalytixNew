"""
Report Export Service - Phase 3 Task 2

Handles exporting reports to multiple formats:
- PDF (via WeasyPrint or ReportLab)
- Excel (via openpyxl)
- CSV (via csv module)
"""

import csv
from io import BytesIO, StringIO
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Tuple
import logging

logger = logging.getLogger(__name__)


class ReportExportService:
    """
    Service for exporting financial reports to various formats.
    
    Supports:
    - CSV format (streamable, lightweight)
    - Excel format (formatted, with calculations)
    - PDF format (print-ready, styled)
    """
    @staticmethod
    def _human_title(report_type: Any, fallback: str = "Report") -> str:
        text = report_type or fallback
        return str(text).replace("_", " ").title()

    @staticmethod
    def _format_timestamp(value: Any) -> str:
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value or "")

    @staticmethod
    def _decimal_to_str(value: Any) -> str:
        if value is None or value == "":
            return "0.00"
        if isinstance(value, Decimal):
            return f"{value:.2f}"
        try:
            return f"{Decimal(str(value)):.2f}"
        except (InvalidOperation, ValueError, TypeError):
            return str(value)

    @staticmethod
    def _decimal_to_float(value: Any) -> float:
        if value is None or value == "":
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        try:
            return float(Decimal(str(value)))
        except (InvalidOperation, ValueError, TypeError):
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0.0

    @staticmethod
    def _prepare_generic_dataset(report_data: Dict[str, Any]) -> Tuple[List[str], List[List[Any]]]:
        rows = report_data.get("rows") or []
        columns = report_data.get("columns")

        if not columns and rows:
            first_row = rows[0]
            if isinstance(first_row, dict):
                columns = list(first_row.keys())
            else:
                columns = [f"Column {idx + 1}" for idx, _ in enumerate(first_row)]

        prepared_rows: List[List[Any]] = []
        if rows:
            if isinstance(rows[0], dict):
                for row in rows:
                    prepared_rows.append([row.get(column) for column in columns])
            else:
                for row in rows:
                    prepared_rows.append(list(row))

        return list(columns or []), prepared_rows

    @staticmethod
    def to_csv(report_data: Dict[str, Any]) -> Tuple[BytesIO, str]:
        """
        Export report to CSV format.
        
        Args:
            report_data: Report data dictionary from ReportService
            
        Returns:
            Tuple of (BytesIO buffer, filename)
        """
        logger.info("Exporting report to CSV: %s", report_data.get("report_type"))
        
        output = StringIO()
        writer = csv.writer(output)

        report_type = report_data.get("report_type") or "report"
        title = report_data.get("name") or ReportExportService._human_title(report_type)

        # Write header
        writer.writerow([f"{title}"])
        organization = report_data.get("organization")
        if organization:
            writer.writerow([f"Organization: {organization}"])
        if "as_of_date" in report_data and report_data["as_of_date"]:
            writer.writerow([f"As of: {report_data['as_of_date']}"])
        elif report_data.get("period"):
            writer.writerow([f"Period: {report_data['period']}"])
        generated = ReportExportService._format_timestamp(report_data.get("generated_at"))
        if generated:
            writer.writerow([f"Generated: {generated}"])
        writer.writerow([])  # Blank line
        
        # Write data based on report type
        handled = True
        if report_type == "general_ledger":
            ReportExportService._export_ledger_csv(writer, report_data)
        elif report_type == "trial_balance":
            ReportExportService._export_trial_balance_csv(writer, report_data)
        elif report_type == "profit_loss":
            ReportExportService._export_pl_csv(writer, report_data)
        elif report_type == "balance_sheet":
            ReportExportService._export_bs_csv(writer, report_data)
        elif report_type == "cash_flow":
            ReportExportService._export_cf_csv(writer, report_data)
        elif report_type == "ar_aging":
            ReportExportService._export_ar_aging_csv(writer, report_data)
        elif report_type == "ap_aging":
            ReportExportService._export_ap_aging_csv(writer, report_data)
        else:
            handled = False

        if not handled or not output.getvalue().strip():
            ReportExportService._export_generic_csv(writer, report_data)
        
        # Convert to bytes
        csv_bytes = BytesIO(output.getvalue().encode("utf-8"))
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return csv_bytes, filename
    
    @staticmethod
    def to_excel(report_data: Dict[str, Any]) -> Tuple[BytesIO, str]:
        """
        Export report to Excel format.
        
        Requires: openpyxl package
        
        Args:
            report_data: Report data dictionary
            
        Returns:
            Tuple of (BytesIO buffer, filename)
        """
        logger.info("Exporting report to Excel: %s", report_data.get("report_type"))
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Define styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header
        row = 1
        ws.cell(row, 1, f"{report_data.get('name') or ReportExportService._human_title(report_data.get('report_type'))} Report")
        ws.cell(row, 1).font = title_font
        row += 1
        
        organization = report_data.get("organization")
        if organization:
            ws.cell(row, 1, f"Organization: {organization}")
            row += 1
        
        if report_data.get("as_of_date"):
            ws.cell(row, 1, f"As of: {report_data['as_of_date']}")
            row += 1
        elif report_data.get("period"):
            ws.cell(row, 1, f"Period: {report_data['period']}")
            row += 1
        
        generated = ReportExportService._format_timestamp(report_data.get("generated_at"))
        if generated:
            ws.cell(row, 1, f"Generated: {generated}")
            row += 1
        row += 1
        
        # Write report data based on type
        report_type = report_data.get("report_type")
        handled = True
        if report_type == "general_ledger":
            row = ReportExportService._export_ledger_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "trial_balance":
            row = ReportExportService._export_trial_balance_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "profit_loss":
            row = ReportExportService._export_pl_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "balance_sheet":
            row = ReportExportService._export_bs_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "cash_flow":
            row = ReportExportService._export_cf_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "ar_aging":
            row = ReportExportService._export_ar_aging_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        elif report_type == "ap_aging":
            row = ReportExportService._export_ap_aging_excel(ws, report_data, row, header_fill, header_font_white, border, total_fill, total_font)
        else:
            handled = False

        if not handled:
            row = ReportExportService._export_generic_excel(ws, report_data, row, header_fill, header_font_white, border)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to bytes
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return excel_bytes, filename
    
    @staticmethod
    def to_pdf(report_data: Dict[str, Any]) -> Tuple[BytesIO, str]:
        """
        Export report to PDF format.
        
        Requires: WeasyPrint or ReportLab package
        
        Args:
            report_data: Report data dictionary
            
        Returns:
            Tuple of (BytesIO buffer, filename)
        """
        logger.info("Exporting report to PDF: %s", report_data.get("report_type"))
        
        try:
            from weasyprint import HTML
            from io import BytesIO
        except ImportError:
            logger.error("WeasyPrint not installed. Install with: pip install weasyprint")
            raise ImportError("WeasyPrint is required for PDF export")
        
        # Generate HTML
        html_content = ReportExportService._generate_pdf_html(report_data)
        
        # Convert to PDF
        pdf_bytes = BytesIO()
        HTML(string=html_content).write_pdf(pdf_bytes)
        pdf_bytes.seek(0)
        
        report_type = report_data.get("report_type") or "report"
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return pdf_bytes, filename
    
    # CSV Export Helpers
    
    @staticmethod
    def _export_ledger_csv(writer, report_data: Dict) -> None:
        """Export General Ledger to CSV."""
        writer.writerow(["Date", "Account", "Journal #", "Reference", "Description", "Debit", "Credit", "Balance"])

        for line in report_data.get("lines", []):
            account_bits = [line.get("account_code") or "", line.get("account_name") or ""]
            account_label = " ".join(bit for bit in account_bits if bit).strip()
            writer.writerow(
                [
                    line.get("date") or "",
                    account_label,
                    line.get("journal_no") or "",
                    line.get("reference") or "",
                    line.get("description") or "",
                    ReportExportService._decimal_to_str(line.get("debit")),
                    ReportExportService._decimal_to_str(line.get("credit")),
                    ReportExportService._decimal_to_str(line.get("running_balance") or line.get("balance")),
                ]
            )

        totals = report_data.get("totals") or {}
        writer.writerow([])
        writer.writerow(["Opening Balance", ReportExportService._decimal_to_str(totals.get("opening_balance"))])
        writer.writerow(["Total Debit", ReportExportService._decimal_to_str(totals.get("total_debit"))])
        writer.writerow(["Total Credit", ReportExportService._decimal_to_str(totals.get("total_credit"))])
        writer.writerow(["Ending Balance", ReportExportService._decimal_to_str(totals.get("ending_balance"))])
    
    @staticmethod
    def _export_trial_balance_csv(writer, report_data: Dict) -> None:
        """Export Trial Balance to CSV."""
        writer.writerow(["Account Code", "Account Name", "Type", "Debit", "Credit"])

        for line in report_data.get("lines", []):
            writer.writerow(
                [
                    line.get("account_code") or "",
                    line.get("account_name") or "",
                    line.get("account_type") or "",
                    ReportExportService._decimal_to_str(line.get("debit_balance")),
                    ReportExportService._decimal_to_str(line.get("credit_balance")),
                ]
            )

        totals = report_data.get("totals") or {}
        writer.writerow([])
        writer.writerow(
            [
                "TOTALS",
                "",
                "",
                ReportExportService._decimal_to_str(totals.get("total_debits")),
                ReportExportService._decimal_to_str(totals.get("total_credits")),
            ]
        )
        writer.writerow(
            ["Difference", "", "", "", ReportExportService._decimal_to_str(totals.get("difference"))]
        )
        writer.writerow(
            ["Balanced", "Yes" if report_data.get("is_balanced") else "No", "", "", ""]
        )
        
        if report_data.get('is_balanced'):
            writer.writerow(['BALANCED: YES'])
        else:
            writer.writerow(['BALANCED: NO - OUT OF BALANCE'])
    
    @staticmethod
    def _export_pl_csv(writer, report_data: Dict) -> None:
        """Export P&L to CSV."""
        sections = report_data.get("sections", [])
        for section in sections:
            writer.writerow([])
            writer.writerow([section.get("category", "").upper()])
            writer.writerow(["Account Code", "Account Name", "Debit", "Credit", "Net"])
            for account in section.get("accounts", []):
                writer.writerow(
                    [
                        account.get("account_code") or "",
                        account.get("account_name") or "",
                        ReportExportService._decimal_to_str(account.get("debit")),
                        ReportExportService._decimal_to_str(account.get("credit")),
                        ReportExportService._decimal_to_str(account.get("net")),
                    ]
                )
            writer.writerow(
                [
                    "Subtotal",
                    "",
                    "",
                    "",
                    ReportExportService._decimal_to_str(section.get("total")),
                ]
            )

        totals = report_data.get("totals") or {}
        if sections:
            writer.writerow([])
        writer.writerow(["Total Income", "", "", "", ReportExportService._decimal_to_str(totals.get("total_income"))])
        writer.writerow(
            ["Total Expense", "", "", "", ReportExportService._decimal_to_str(totals.get("total_expense"))]
        )
        writer.writerow(["Net Profit", "", "", "", ReportExportService._decimal_to_str(totals.get("net_profit"))])
    
    @staticmethod
    def _export_bs_csv(writer, report_data: Dict) -> None:
        """Export Balance Sheet to CSV."""
        lines = report_data.get("lines", [])
        grouped: "OrderedDict[str, OrderedDict[str, List[Dict[str, Any]]]]" = OrderedDict()

        for entry in lines:
            nature = entry.get("line_type") or entry.get("nature") or "Other"
            category = entry.get("category") or nature
            nature_key = ReportExportService._human_title(nature)
            category_key = ReportExportService._human_title(category)
            category_map = grouped.setdefault(nature_key, OrderedDict())
            bucket = category_map.setdefault(category_key, [])
            bucket.append(entry)

        for nature, category_map in grouped.items():
            writer.writerow([])
            writer.writerow([nature])
            writer.writerow(["Account Code", "Account Name", "Category", "Amount"])
            nature_total = 0.0
            for category, entries in category_map.items():
                for item in entries:
                    amount = ReportExportService._decimal_to_float(item.get("amount"))
                    nature_total += amount
                    writer.writerow(
                        [
                            item.get("account_code") or "",
                            item.get("account_name") or "",
                            category,
                            ReportExportService._decimal_to_str(item.get("amount")),
                        ]
                    )
            writer.writerow(["Subtotal", "", "", ReportExportService._decimal_to_str(nature_total)])

        totals = report_data.get("totals") or {}
        if grouped:
            writer.writerow([])
        writer.writerow(["Total Assets", "", "", ReportExportService._decimal_to_str(totals.get("total_assets"))])
        writer.writerow(
            ["Total Liabilities", "", "", ReportExportService._decimal_to_str(totals.get("total_liabilities"))]
        )
        writer.writerow(["Total Equity", "", "", ReportExportService._decimal_to_str(totals.get("total_equity"))])
        writer.writerow(
            [
                "Liabilities + Equity",
                "",
                "",
                ReportExportService._decimal_to_str(totals.get("total_liabilities_equity")),
            ]
        )
        writer.writerow(["Difference", "", "", ReportExportService._decimal_to_str(totals.get("difference"))])
    
    @staticmethod
    def _export_cf_csv(writer, report_data: Dict) -> None:
        """Export Cash Flow to CSV."""
        categories = report_data.get("categories", [])
        for bucket in categories:
            writer.writerow([])
            writer.writerow([bucket.get("category", "").upper()])
            writer.writerow(["Account Code", "Account Name", "Amount"])
            for account in bucket.get("accounts", []):
                writer.writerow(
                    [
                        account.get("account_code") or "",
                        account.get("account_name") or "",
                        ReportExportService._decimal_to_str(account.get("amount")),
                    ]
                )
            writer.writerow(
                ["Subtotal", "", ReportExportService._decimal_to_str(bucket.get("total"))]
            )

        totals = report_data.get("totals") or {}
        if categories:
            writer.writerow([])
        writer.writerow(["Net Cash Movement", "", ReportExportService._decimal_to_str(totals.get("net_change"))])
    
    @staticmethod
    def _export_ar_aging_csv(writer, report_data: Dict) -> None:
        """Export A/R Aging to CSV."""
        writer.writerow(["Account Code", "Account Name", "Bucket", "Balance"])

        for line in report_data.get("lines", []):
            writer.writerow(
                [
                    line.get("account_code") or "",
                    line.get("account_name") or "",
                    line.get("bucket") or "",
                    ReportExportService._decimal_to_str(line.get("balance")),
                ]
            )

        writer.writerow([])
        writer.writerow(["AGING SUMMARY"])
        for bucket in report_data.get("aging_summary", []):
            writer.writerow(
                [bucket.get("bucket") or "", ReportExportService._decimal_to_str(bucket.get("balance"))]
            )
        writer.writerow(["TOTAL", ReportExportService._decimal_to_str(report_data.get("total"))])

    @staticmethod
    def _export_ap_aging_csv(writer, report_data: Dict) -> None:
        """Export A/P Aging to CSV."""
        writer.writerow(["Vendor", "Bucket", "Balance"])

        for line in report_data.get("lines", []):
            writer.writerow(
                [
                    line.get("vendor_name") or "",
                    line.get("bucket") or "",
                    ReportExportService._decimal_to_str(line.get("balance")),
                ]
            )

        writer.writerow([])
        writer.writerow(["AGING SUMMARY"])
        for bucket in report_data.get("aging_summary", []):
            writer.writerow(
                [bucket.get("bucket") or "", ReportExportService._decimal_to_str(bucket.get("balance"))]
            )
        writer.writerow(["TOTAL", ReportExportService._decimal_to_str(report_data.get("total"))])

    @staticmethod
    def _export_generic_csv(writer, report_data: Dict) -> None:
        """Fallback CSV export for dynamic/custom reports."""
        columns, rows = ReportExportService._prepare_generic_dataset(report_data)
        if columns:
            writer.writerow(columns)
        for record in rows:
            writer.writerow(record)
    
    # Excel Export Helpers
    
    @staticmethod
    def _export_ledger_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export General Ledger to Excel sheet."""
        # Headers
        headers = ["Date", "Account", "Journal #", "Reference", "Description", "Debit", "Credit", "Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data
        for line in report_data.get("lines", []):
            account_bits = [line.get("account_code") or "", line.get("account_name") or ""]
            account_label = " ".join(bit for bit in account_bits if bit).strip()
            ws.cell(row, 1, line.get("date"))
            ws.cell(row, 2, account_label)
            ws.cell(row, 3, line.get("journal_no"))
            ws.cell(row, 4, line.get("reference"))
            ws.cell(row, 5, line.get("description"))
            ws.cell(row, 6, ReportExportService._decimal_to_float(line.get("debit")))
            ws.cell(row, 7, ReportExportService._decimal_to_float(line.get("credit")))
            ws.cell(row, 8, ReportExportService._decimal_to_float(line.get("running_balance") or line.get("balance")))
            row += 1
        
        # Totals
        row += 1
        totals = report_data.get("totals") or {}
        ws.cell(row, 1, "Opening Balance")
        ws.cell(row, 2, ReportExportService._decimal_to_float(totals.get("opening_balance")))
        ws.cell(row + 1, 1, "Total Debit")
        ws.cell(row + 1, 2, ReportExportService._decimal_to_float(totals.get("total_debit")))
        ws.cell(row + 2, 1, "Total Credit")
        ws.cell(row + 2, 2, ReportExportService._decimal_to_float(totals.get("total_credit")))
        ws.cell(row + 3, 1, "Ending Balance")
        ws.cell(row + 3, 2, ReportExportService._decimal_to_float(totals.get("ending_balance")))

        for offset in range(0, 4):
            for col in range(1, 3):
                ws.cell(row + offset, col).fill = total_fill
                ws.cell(row + offset, col).font = total_font

        return row + 4
    
    @staticmethod
    def _export_trial_balance_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export Trial Balance to Excel sheet."""
        headers = ["Account Code", "Account Name", "Type", "Debit", "Credit"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = border
        row += 1
        
        for line in report_data.get("lines", []):
            ws.cell(row, 1, line.get("account_code"))
            ws.cell(row, 2, line.get("account_name"))
            ws.cell(row, 3, line.get("account_type"))
            ws.cell(row, 4, ReportExportService._decimal_to_float(line.get("debit_balance")))
            ws.cell(row, 5, ReportExportService._decimal_to_float(line.get("credit_balance")))
            row += 1
        
        row += 1
        totals = report_data.get("totals") or {}
        ws.cell(row, 1, "TOTALS")
        ws.cell(row, 4, ReportExportService._decimal_to_float(totals.get("total_debits")))
        ws.cell(row, 5, ReportExportService._decimal_to_float(totals.get("total_credits")))
        for col in range(1, 6):
            ws.cell(row, col).fill = total_fill
            ws.cell(row, col).font = total_font
        row += 1
        ws.cell(row, 1, "Difference")
        ws.cell(row, 5, ReportExportService._decimal_to_float(totals.get("difference")))
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, 1).font = total_font
        ws.cell(row, 5).fill = total_fill
        ws.cell(row, 5).font = total_font
        row += 1
        ws.cell(row, 1, "Balanced")
        ws.cell(row, 2, "Yes" if report_data.get("is_balanced") else "No")
        return row + 1
    
    @staticmethod
    def _export_pl_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export P&L to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        sections = report_data.get("sections", [])
        for section in sections:
            ws.cell(row, 1, section.get("category", "").upper())
            ws.cell(row, 1).font = Font(bold=True, size=11)
            row += 1

            headers = ["Account Code", "Account Name", "Debit", "Credit", "Net"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for account in section.get("accounts", []):
                ws.cell(row, 1, account.get("account_code"))
                ws.cell(row, 2, account.get("account_name"))
                ws.cell(row, 3, ReportExportService._decimal_to_float(account.get("debit")))
                ws.cell(row, 4, ReportExportService._decimal_to_float(account.get("credit")))
                ws.cell(row, 5, ReportExportService._decimal_to_float(account.get("net")))
                row += 1

            ws.cell(row, 1, "Subtotal")
            ws.cell(row, 5, ReportExportService._decimal_to_float(section.get("total")))
            ws.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 5).fill = total_fill
            ws.cell(row, 5).font = total_font
            row += 2

        totals = report_data.get("totals") or {}
        ws.cell(row, 1, "Total Income")
        ws.cell(row, 2, ReportExportService._decimal_to_float(totals.get("total_income")))
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, 1).font = total_font
        ws.cell(row, 2).fill = total_fill
        ws.cell(row, 2).font = total_font
        row += 1

        ws.cell(row, 1, "Total Expense")
        ws.cell(row, 2, ReportExportService._decimal_to_float(totals.get("total_expense")))
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, 1).font = total_font
        ws.cell(row, 2).fill = total_fill
        ws.cell(row, 2).font = total_font
        row += 1

        ws.cell(row, 1, "Net Profit")
        ws.cell(row, 2, ReportExportService._decimal_to_float(totals.get("net_profit")))
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, 1).font = total_font
        ws.cell(row, 2).fill = total_fill
        ws.cell(row, 2).font = total_font
        row += 1

        return row
    
    @staticmethod
    def _export_bs_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export Balance Sheet to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        lines = report_data.get("lines", [])
        grouped: "OrderedDict[str, OrderedDict[str, List[Dict[str, Any]]]]" = OrderedDict()

        for entry in lines:
            nature = entry.get("line_type") or entry.get("nature") or "Other"
            category = entry.get("category") or nature
            nature_key = ReportExportService._human_title(nature)
            category_key = ReportExportService._human_title(category)
            category_map = grouped.setdefault(nature_key, OrderedDict())
            bucket = category_map.setdefault(category_key, [])
            bucket.append(entry)

        for nature, category_map in grouped.items():
            ws.cell(row, 1, nature)
            ws.cell(row, 1).font = Font(bold=True, size=11)
            row += 1

            headers = ["Account Code", "Account Name", "Category", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            nature_total = 0.0
            for category, entries in category_map.items():
                for item in entries:
                    nature_total += ReportExportService._decimal_to_float(item.get("amount"))
                    ws.cell(row, 1, item.get("account_code"))
                    ws.cell(row, 2, item.get("account_name"))
                    ws.cell(row, 3, category)
                    ws.cell(row, 4, ReportExportService._decimal_to_float(item.get("amount")))
                    row += 1

            ws.cell(row, 1, "Subtotal")
            ws.cell(row, 4, nature_total)
            ws.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 4).fill = total_fill
            ws.cell(row, 4).font = total_font
            row += 2

        totals = report_data.get("totals") or {}
        summary_rows = [
            ("Total Assets", totals.get("total_assets")),
            ("Total Liabilities", totals.get("total_liabilities")),
            ("Total Equity", totals.get("total_equity")),
            ("Liabilities + Equity", totals.get("total_liabilities_equity")),
            ("Difference", totals.get("difference")),
        ]
        for label, value in summary_rows:
            ws.cell(row, 1, label)
            ws.cell(row, 2, ReportExportService._decimal_to_float(value))
            ws.cell(row, 1).fill = total_fill
            ws.cell(row, 1).font = total_font
            ws.cell(row, 2).fill = total_fill
            ws.cell(row, 2).font = total_font
            row += 1

        return row
    
    @staticmethod
    def _export_cf_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export Cash Flow to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        categories = report_data.get("categories", [])
        for bucket in categories:
            ws.cell(row, 1, bucket.get("category", "").upper())
            ws.cell(row, 1).font = Font(bold=True, size=11)
            row += 1

            headers = ["Account Code", "Account Name", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for account in bucket.get("accounts", []):
                ws.cell(row, 1, account.get("account_code"))
                ws.cell(row, 2, account.get("account_name"))
                ws.cell(row, 3, ReportExportService._decimal_to_float(account.get("amount")))
                row += 1

            ws.cell(row, 1, "Subtotal")
            ws.cell(row, 3, ReportExportService._decimal_to_float(bucket.get("total")))
            ws.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 3).fill = total_fill
            ws.cell(row, 3).font = total_font
            row += 2

        totals = report_data.get("totals") or {}
        ws.cell(row, 1, "Net Cash Movement")
        ws.cell(row, 2, ReportExportService._decimal_to_float(totals.get("net_change")))
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, 1).font = total_font
        ws.cell(row, 2).fill = total_fill
        ws.cell(row, 2).font = total_font
        row += 1

        return row
    
    @staticmethod
    def _export_ar_aging_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export A/R Aging to Excel sheet."""
        from openpyxl.styles import Font

        headers = ["Account Code", "Account Name", "Bucket", "Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = border
        row += 1

        for line in report_data.get("lines", []):
            ws.cell(row, 1, line.get("account_code"))
            ws.cell(row, 2, line.get("account_name"))
            ws.cell(row, 3, line.get("bucket"))
            ws.cell(row, 4, ReportExportService._decimal_to_float(line.get("balance")))
            row += 1

        row += 2
        ws.cell(row, 1, "AGING SUMMARY")
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1

        for bucket in report_data.get("aging_summary", []):
            ws.cell(row, 1, bucket.get("bucket"))
            ws.cell(row, 2, ReportExportService._decimal_to_float(bucket.get("balance")))
            row += 1

        ws.cell(row, 1, "TOTAL")
        ws.cell(row, 2, ReportExportService._decimal_to_float(report_data.get("total")))
        for col in range(1, 3):
            ws.cell(row, col).fill = total_fill
            ws.cell(row, col).font = Font(bold=True)

        return row + 1

    @staticmethod
    def _export_ap_aging_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border, total_fill, total_font) -> int:
        """Export A/P Aging to Excel sheet."""
        from openpyxl.styles import Font

        headers = ["Vendor", "Bucket", "Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = border
        row += 1

        for line in report_data.get("lines", []):
            ws.cell(row, 1, line.get("vendor_name"))
            ws.cell(row, 2, line.get("bucket"))
            ws.cell(row, 3, ReportExportService._decimal_to_float(line.get("balance")))
            row += 1

        row += 2
        ws.cell(row, 1, "AGING SUMMARY")
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1

        for bucket in report_data.get("aging_summary", []):
            ws.cell(row, 1, bucket.get("bucket"))
            ws.cell(row, 2, ReportExportService._decimal_to_float(bucket.get("balance")))
            row += 1

        ws.cell(row, 1, "TOTAL")
        ws.cell(row, 2, ReportExportService._decimal_to_float(report_data.get("total")))
        for col in range(1, 3):
            ws.cell(row, col).fill = total_fill
            ws.cell(row, col).font = Font(bold=True)

        return row + 1

    # PDF Generation Helper

    @staticmethod
    def _export_generic_excel(ws, report_data: Dict, row: int, header_fill, header_font_white, border) -> int:
        """Fallback Excel export for custom datasets."""
        columns, rows = ReportExportService._prepare_generic_dataset(report_data)
        if columns:
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row, col, header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = border
            row += 1

        for record in rows:
            for col, value in enumerate(record, 1):
                ws.cell(row, col, value)
            row += 1

        return row

    @staticmethod
    def _generate_pdf_html(report_data: Dict) -> str:
        """Generate HTML for PDF rendering."""
        report_type = report_data.get("report_type")
        title = report_data.get("name") or ReportExportService._human_title(report_type)

        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
                .title {{ font-size: 18px; font-weight: bold; margin-bottom: 10px; }}
                .meta {{ font-size: 12px; color: #666; margin-bottom: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th {{ background-color: #4472C4; color: white; padding: 8px; text-align: left; border: 1px solid #999; }}
                td {{ padding: 8px; border: 1px solid #999; }}
                .total {{ background-color: #E7E6E6; font-weight: bold; }}
                .number {{ text-align: right; }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{title}</div>
        """

        organization = report_data.get("organization")
        if organization:
            html += f"<div class='meta'>Organization: {organization}</div>"

        if report_data.get("as_of_date"):
            html += f"<div class='meta'>As of: {report_data['as_of_date']}</div>"
        elif report_data.get("period"):
            html += f"<div class='meta'>Period: {report_data['period']}</div>"

        generated = ReportExportService._format_timestamp(report_data.get("generated_at"))
        if generated:
            html += f"<div class='meta'>Generated: {generated}</div>"

        html += "</div>"

        handler_map = {
            "general_ledger": ReportExportService._generate_ledger_html_table,
            "trial_balance": ReportExportService._generate_trial_balance_html_table,
            "profit_loss": ReportExportService._generate_pl_html_table,
            "balance_sheet": ReportExportService._generate_bs_html_table,
            "cash_flow": ReportExportService._generate_cf_html_table,
            "ar_aging": ReportExportService._generate_ar_aging_html_table,
        }

        handler = handler_map.get(report_type)
        if handler:
            html += handler(report_data)
        else:
            html += ReportExportService._generate_generic_html_table(report_data)

        html += "</body></html>"
        return html
    
    @staticmethod
    def _generate_ledger_html_table(report_data: Dict) -> str:
        """Generate HTML table for General Ledger."""
        html = """<table>
            <tr>
                <th>Date</th>
                <th>Account</th>
                <th>Journal #</th>
                <th>Reference</th>
                <th>Description</th>
                <th class="number">Debit</th>
                <th class="number">Credit</th>
                <th class="number">Balance</th>
            </tr>"""

        for line in report_data.get("lines", []):
            account_bits = [line.get("account_code") or "", line.get("account_name") or ""]
            account_label = " ".join(bit for bit in account_bits if bit).strip()
            html += f"""<tr>
                <td>{line.get('date') or ''}</td>
                <td>{account_label}</td>
                <td>{line.get('journal_no') or ''}</td>
                <td>{line.get('reference') or ''}</td>
                <td>{line.get('description') or ''}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('debit'))}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('credit'))}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('running_balance') or line.get('balance'))}</td>
            </tr>"""

        html += "</table>"

        totals = report_data.get("totals") or {}
        html += """<table>
            <tr class="total"><td>Opening Balance</td><td class="number">{opening}</td></tr>
            <tr class="total"><td>Total Debit</td><td class="number">{debit}</td></tr>
            <tr class="total"><td>Total Credit</td><td class="number">{credit}</td></tr>
            <tr class="total"><td>Ending Balance</td><td class="number">{ending}</td></tr>
        </table>""".format(
            opening=ReportExportService._decimal_to_str(totals.get("opening_balance")),
            debit=ReportExportService._decimal_to_str(totals.get("total_debit")),
            credit=ReportExportService._decimal_to_str(totals.get("total_credit")),
            ending=ReportExportService._decimal_to_str(totals.get("ending_balance")),
        )

        return html
    
    @staticmethod
    def _generate_trial_balance_html_table(report_data: Dict) -> str:
        """Generate HTML table for Trial Balance."""
        html = """<table>
            <tr>
                <th>Account Code</th>
                <th>Account Name</th>
                <th>Type</th>
                <th class="number">Debit</th>
                <th class="number">Credit</th>
            </tr>"""
        
        for line in report_data.get("lines", []):
            html += f"""<tr>
                <td>{line.get('account_code') or ''}</td>
                <td>{line.get('account_name') or ''}</td>
                <td>{line.get('account_type') or ''}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('debit_balance'))}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('credit_balance'))}</td>
            </tr>"""
        
        html += f"""<tr class="total">
            <td colspan="3">TOTALS</td>
            <td class="number">{ReportExportService._decimal_to_str(report_data.get('totals', {}).get('total_debits'))}</td>
            <td class="number">{ReportExportService._decimal_to_str(report_data.get('totals', {}).get('total_credits'))}</td>
        </tr></table>"""
        
        totals = report_data.get("totals") or {}
        html += f"<p><strong>Difference:</strong> {ReportExportService._decimal_to_str(totals.get('difference'))}</p>"

        balanced_status = "YES" if report_data.get("is_balanced") else "NO - OUT OF BALANCE"
        html += f"<p><strong>Balanced:</strong> {balanced_status}</p>"
        
        return html
    
    @staticmethod
    def _generate_pl_html_table(report_data: Dict) -> str:
        """Generate HTML table for P&L Statement."""
        html = ""

        for section in report_data.get("sections", []):
            html += "<table>"
            html += f"<tr><th colspan='5'>{section.get('category', '').upper()}</th></tr>"
            html += """
                <tr>
                    <th>Account Code</th>
                    <th>Account Name</th>
                    <th class="number">Debit</th>
                    <th class="number">Credit</th>
                    <th class="number">Net</th>
                </tr>
            """
            for account in section.get("accounts", []):
                html += f"""<tr>
                    <td>{account.get('account_code') or ''}</td>
                    <td>{account.get('account_name') or ''}</td>
                    <td class="number">{ReportExportService._decimal_to_str(account.get('debit'))}</td>
                    <td class="number">{ReportExportService._decimal_to_str(account.get('credit'))}</td>
                    <td class="number">{ReportExportService._decimal_to_str(account.get('net'))}</td>
                </tr>"""
            html += f"""<tr class="total">
                <td colspan="4">Subtotal</td>
                <td class="number">{ReportExportService._decimal_to_str(section.get('total'))}</td>
            </tr></table>"""

        totals = report_data.get("totals") or {}
        html += "<table>"
        html += f"""<tr class="total"><td>Total Income</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_income'))}</td></tr>"""
        html += f"""<tr class="total"><td>Total Expense</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_expense'))}</td></tr>"""
        html += f"""<tr class="total"><td>Net Profit</td><td class="number">{ReportExportService._decimal_to_str(totals.get('net_profit'))}</td></tr>"""
        html += "</table>"
        return html
    
    @staticmethod
    def _generate_bs_html_table(report_data: Dict) -> str:
        """Generate HTML table for Balance Sheet."""
        lines = report_data.get("lines", [])
        grouped: "OrderedDict[str, OrderedDict[str, List[Dict[str, Any]]]]" = OrderedDict()

        for entry in lines:
            nature = entry.get("line_type") or entry.get("nature") or "Other"
            category = entry.get("category") or nature
            nature_key = ReportExportService._human_title(nature)
            category_key = ReportExportService._human_title(category)
            category_map = grouped.setdefault(nature_key, OrderedDict())
            bucket = category_map.setdefault(category_key, [])
            bucket.append(entry)

        html = ""
        for nature, category_map in grouped.items():
            html += "<table>"
            html += f"<tr><th colspan='4'>{nature}</th></tr>"
            html += """
                <tr>
                    <th>Account Code</th>
                    <th>Account Name</th>
                    <th>Category</th>
                    <th class="number">Amount</th>
                </tr>
            """
            nature_total = Decimal("0.00")
            for category, entries in category_map.items():
                for item in entries:
                    raw_amount = item.get("amount")
                    try:
                        amount_decimal = raw_amount if isinstance(raw_amount, Decimal) else Decimal(str(raw_amount or 0))
                    except (InvalidOperation, TypeError):
                        amount_decimal = Decimal("0.00")
                    nature_total += amount_decimal
                    html += f"""<tr>
                        <td>{item.get('account_code') or ''}</td>
                        <td>{item.get('account_name') or ''}</td>
                        <td>{category}</td>
                        <td class="number">{ReportExportService._decimal_to_str(item.get('amount'))}</td>
                    </tr>"""
            html += f"""<tr class="total">
                <td colspan="3">Subtotal</td>
                <td class="number">{ReportExportService._decimal_to_str(nature_total)}</td>
            </tr></table>"""

        totals = report_data.get("totals") or {}
        html += "<table>"
        html += f"""<tr class="total"><td>Total Assets</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_assets'))}</td></tr>"""
        html += f"""<tr class="total"><td>Total Liabilities</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_liabilities'))}</td></tr>"""
        html += f"""<tr class="total"><td>Total Equity</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_equity'))}</td></tr>"""
        html += f"""<tr class="total"><td>Liabilities + Equity</td><td class="number">{ReportExportService._decimal_to_str(totals.get('total_liabilities_equity'))}</td></tr>"""
        html += f"""<tr class="total"><td>Difference</td><td class="number">{ReportExportService._decimal_to_str(totals.get('difference'))}</td></tr>"""
        html += "</table>"

        return html
    
    @staticmethod
    def _generate_cf_html_table(report_data: Dict) -> str:
        """Generate HTML table for Cash Flow."""
        html = ""
        for bucket in report_data.get("categories", []):
            html += "<table>"
            html += f"<tr><th colspan='3'>{bucket.get('category', '').upper()}</th></tr>"
            html += """
                <tr>
                    <th>Account Code</th>
                    <th>Account Name</th>
                    <th class="number">Amount</th>
                </tr>
            """
            for account in bucket.get("accounts", []):
                html += f"""<tr>
                    <td>{account.get('account_code') or ''}</td>
                    <td>{account.get('account_name') or ''}</td>
                    <td class="number">{ReportExportService._decimal_to_str(account.get('amount'))}</td>
                </tr>"""
            html += f"""<tr class="total">
                <td colspan="2">Subtotal</td>
                <td class="number">{ReportExportService._decimal_to_str(bucket.get('total'))}</td>
            </tr></table>"""

        totals = report_data.get("totals") or {}
        html += "<table>"
        html += f"""<tr class="total"><td>Net Cash Movement</td><td class="number">{ReportExportService._decimal_to_str(totals.get('net_change'))}</td></tr>"""
        html += "</table>"
        return html
    
    @staticmethod
    def _generate_ar_aging_html_table(report_data: Dict) -> str:
        """Generate HTML table for A/R Aging."""
        html = """<table>
            <tr>
                <th>Account Code</th>
                <th>Account Name</th>
                <th>Bucket</th>
                <th class="number">Balance</th>
            </tr>"""

        for line in report_data.get("lines", []):
            html += f"""<tr>
                <td>{line.get('account_code') or ''}</td>
                <td>{line.get('account_name') or ''}</td>
                <td>{line.get('bucket') or ''}</td>
                <td class="number">{ReportExportService._decimal_to_str(line.get('balance'))}</td>
            </tr>"""

        html += """</table>
            <h3>Aging Summary</h3>
            <table>
                <tr><th>Bucket</th><th class="number">Amount</th></tr>"""

        for bucket in report_data.get("aging_summary", []):
            html += f"""<tr>
                <td>{bucket.get('bucket') or ''}</td>
                <td class="number">{ReportExportService._decimal_to_str(bucket.get('balance'))}</td>
            </tr>"""

        html += f"""<tr class="total">
                <td><strong>TOTAL</strong></td>
                <td class="number"><strong>{ReportExportService._decimal_to_str(report_data.get('total'))}</strong></td>
            </tr></table>"""

        return html

    @staticmethod
    def _generate_generic_html_table(report_data: Dict) -> str:
        """Fallback HTML rendering for custom datasets."""
        columns, rows = ReportExportService._prepare_generic_dataset(report_data)
        if not columns:
            return "<p>No data available for export.</p>"

        html = "<table><tr>"
        for header in columns:
            html += f"<th>{header}</th>"
        html += "</tr>"

        for record in rows:
            html += "<tr>"
            for value in record:
                html += f"<td>{value if value is not None else ''}</td>"
            html += "</tr>"
        html += "</table>"
        return html
