"""
Batch Import/Export Service - Phase 3 Task 3

Comprehensive import/export system for financial data:
- Excel template import with validation
- CSV import with error handling
- Duplicate detection and conflict resolution
- Import history and audit logging
- Progress tracking and batch operations
"""

from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils.translation import gettext as _
from django.contrib.auth import get_user_model
from decimal import Decimal
from datetime import datetime, date
import csv
import logging
from io import StringIO, BytesIO
from typing import Dict, List, Tuple, Optional, TYPE_CHECKING, Any

import openpyxl
from openpyxl.utils import get_column_letter

from accounting.models import Account, Journal, JournalLine, JournalType, Organization

if TYPE_CHECKING:
    from django.contrib.auth.models import AbstractBaseUser
    
logger = logging.getLogger(__name__)
User = get_user_model()


class ImportTemplate:
    """Standard import template definition."""
    
    HEADERS = [
        'Date',
        'Journal Type',
        'Reference',
        'Description',
        'Account Code',
        'Account Name',
        'Debit',
        'Credit',
        'Department',
        'Project',
        'Cost Center'
    ]
    
    REQUIRED_FIELDS = ['Date', 'Account Code', 'Journal Type']
    
    @staticmethod
    def create_excel_template(organization: Organization) -> BytesIO:
        """Create Excel template for import."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal Entries"
        
        # Add headers
        for col, header in enumerate(ImportTemplate.HEADERS, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 15  # Journal Type
        ws.column_dimensions['C'].width = 15  # Reference
        ws.column_dimensions['D'].width = 30  # Description
        ws.column_dimensions['E'].width = 12  # Account Code
        ws.column_dimensions['F'].width = 25  # Account Name
        ws.column_dimensions['G'].width = 12  # Debit
        ws.column_dimensions['H'].width = 12  # Credit
        
        # Add data validation example
        ws['A2'].value = f"{date.today().year}-01-01"
        ws['B2'].value = "GJ"
        ws['C2'].value = "REF001"
        ws['D2'].value = "Sample transaction"
        ws['E2'].value = "1000"
        ws['F2'].value = "Cash"
        ws['G2'].value = 1000.00
        ws['H2'].value = 0.00
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ImportValidationError(Exception):
    """Raised when import validation fails."""
    pass


class DuplicateDetector:
    """Detects duplicate journal entries."""
    
    @staticmethod
    def check_duplicate(
        organization: Organization,
        reference: str,
        date: date,
        amount: Decimal
    ) -> Optional[Journal]:
        """Check if similar journal exists."""
        # Search for journals with same reference and date within tolerance
        similar = Journal.objects.filter(
            organization=organization,
            reference=reference,
            date=date
        ).first()
        
        if similar:
            return similar
        
        return None
    
    @staticmethod
    def check_conflicts(
        organization: Organization,
        proposed_data: Dict
    ) -> List[str]:
        """Check for data conflicts."""
        conflicts = []
        
        # Check account exists
        try:
            Account.objects.get(
                organization=organization,
                code=proposed_data['account_code']
            )
        except Account.DoesNotExist:
            conflicts.append(f"Account {proposed_data['account_code']} not found")
        
        # Check journal type exists
        try:
            JournalType.objects.get(code=proposed_data['journal_type'])
        except JournalType.DoesNotExist:
            conflicts.append(f"Journal Type {proposed_data['journal_type']} not found")
        
        # Validate amounts
        if proposed_data.get('debit', 0) < 0 or proposed_data.get('credit', 0) < 0:
            conflicts.append("Debit/Credit amounts must be positive")
        
        if proposed_data.get('debit', 0) > 0 and proposed_data.get('credit', 0) > 0:
            conflicts.append("Cannot have both debit and credit in same line")
        
        return conflicts


class ImportService:
    """Main import service for batch operations."""
    
    def __init__(self, organization: Organization, user: Any):
        """Initialize import service with organization and user."""
        self.organization = organization
        self.user = user
        self.import_id = f"import_{datetime.now().timestamp()}"
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        self.skipped_count = 0
        self.duplicate_count = 0
    
    def import_excel(self, file_path: str, skip_duplicates: bool = True) -> Dict:
        """
        Import from Excel file.
        
        Args:
            file_path: Path to Excel file
            skip_duplicates: Skip duplicate entries
            
        Returns:
            Import result with statistics
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                row_data = self._parse_excel_row(row, row_idx)
                if row_data:
                    rows.append(row_data)
            
            return self._process_rows(rows, skip_duplicates)
        
        except Exception as e:
            logger.exception(f"Excel import error: {e}")
            self.errors.append(f"Excel import failed: {str(e)}")
            return self._get_result()
    
    def import_csv(self, file_content: str, skip_duplicates: bool = True) -> Dict:
        """
        Import from CSV content.
        
        Args:
            file_content: CSV file content
            skip_duplicates: Skip duplicate entries
            
        Returns:
            Import result with statistics
        """
        try:
            csv_reader = csv.DictReader(StringIO(file_content))
            
            rows = []
            for row_idx, row in enumerate(csv_reader, start=2):
                row_data = self._parse_csv_row(row, row_idx)
                if row_data:
                    rows.append(row_data)
            
            return self._process_rows(rows, skip_duplicates)
        
        except Exception as e:
            logger.exception(f"CSV import error: {e}")
            self.errors.append(f"CSV import failed: {str(e)}")
            return self._get_result()
    
    def _parse_excel_row(self, row: Tuple, row_idx: int) -> Optional[Dict]:
        """Parse Excel row to data dict."""
        try:
            return {
                'date': row[0],
                'journal_type': row[1],
                'reference': row[2],
                'description': row[3],
                'account_code': row[4],
                'account_name': row[5],
                'debit': row[6],
                'credit': row[7],
                'department': row[8],
                'project': row[9],
                'cost_center': row[10],
                'row_number': row_idx
            }
        except Exception as e:
            self.errors.append(f"Row {row_idx}: Parse error - {str(e)}")
            return None
    
    def _parse_csv_row(self, row: Dict, row_idx: int) -> Optional[Dict]:
        """Parse CSV row to data dict."""
        try:
            return {
                'date': row.get('Date'),
                'journal_type': row.get('Journal Type'),
                'reference': row.get('Reference'),
                'description': row.get('Description'),
                'account_code': row.get('Account Code'),
                'account_name': row.get('Account Name'),
                'debit': float(row.get('Debit', 0)) if row.get('Debit') else 0,
                'credit': float(row.get('Credit', 0)) if row.get('Credit') else 0,
                'department': row.get('Department'),
                'project': row.get('Project'),
                'cost_center': row.get('Cost Center'),
                'row_number': row_idx
            }
        except Exception as e:
            self.errors.append(f"Row {row_idx}: Parse error - {str(e)}")
            return None
    
    def _process_rows(self, rows: List[Dict], skip_duplicates: bool) -> Dict:
        """Process and validate rows."""
        # Group by journal transaction
        journal_groups = {}
        
        for row in rows:
            try:
                # Validate required fields
                self._validate_required_fields(row)
                
                # Check conflicts
                conflicts = DuplicateDetector.check_conflicts(self.organization, row)
                if conflicts:
                    self.errors.append(f"Row {row['row_number']}: {', '.join(conflicts)}")
                    self.skipped_count += 1
                    continue
                
                # Check duplicates
                duplicate = DuplicateDetector.check_duplicate(
                    self.organization,
                    row.get('reference', ''),
                    row.get('date'),
                    Decimal(str(row.get('debit', 0) or row.get('credit', 0)))
                )
                
                if duplicate:
                    if skip_duplicates:
                        self.duplicate_count += 1
                        self.warnings.append(f"Row {row['row_number']}: Duplicate skipped")
                        continue
                
                # Group rows by journal reference
                ref_key = (row.get('reference'), row.get('date'))
                if ref_key not in journal_groups:
                    journal_groups[ref_key] = []
                journal_groups[ref_key].append(row)
            
            except Exception as e:
                self.errors.append(f"Row {row['row_number']}: {str(e)}")
                self.skipped_count += 1
        
        # Import journals
        with transaction.atomic():
            for journal_data in journal_groups.values():
                try:
                    self._create_journal_from_rows(journal_data)
                    self.imported_count += 1
                except Exception as e:
                    self.errors.append(f"Journal import failed: {str(e)}")
                    self.skipped_count += 1
        
        return self._get_result()
    
    def _validate_required_fields(self, row: Dict):
        """Validate required fields present."""
        for field in ImportTemplate.REQUIRED_FIELDS:
            if not row.get(field):
                raise ValidationError(f"Required field '{field}' is missing")
    
    def _create_journal_from_rows(self, rows: List[Dict]):
        """Create journal and lines from rows."""
        first_row = rows[0]
        
        # Parse date
        date_value = first_row['date']
        if isinstance(date_value, str):
            date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
        
        # Get journal type
        journal_type = JournalType.objects.get(code=first_row['journal_type'])
        
        # Create journal
        journal = Journal.objects.create(
            organization=self.organization,
            journal_type=journal_type,
            date=date_value,
            reference=first_row.get('reference', ''),
            description=first_row.get('description', ''),
            status='Draft',
            created_by=self.user
        )
        
        # Create journal lines
        for row in rows:
            account = Account.objects.get(
                organization=self.organization,
                code=row['account_code']
            )
            
            JournalLine.objects.create(
                journal=journal,
                account=account,
                debit=Decimal(str(row.get('debit', 0) or 0)),
                credit=Decimal(str(row.get('credit', 0) or 0)),
                description=row.get('description', '')
            )
        
        logger.info(f"Imported journal {journal.id} with {len(rows)} lines")
    
    def _get_result(self) -> Dict:
        """Get import result summary."""
        return {
            'import_id': self.import_id,
            'imported_count': self.imported_count,
            'skipped_count': self.skipped_count,
            'duplicate_count': self.duplicate_count,
            'total_processed': self.imported_count + self.skipped_count + self.duplicate_count,
            'errors': self.errors,
            'warnings': self.warnings,
            'success': len(self.errors) == 0,
            'timestamp': datetime.now()
        }


class ExportService:
    """Service for exporting journals to various formats."""
    
    @staticmethod
    def export_journals_to_excel(journals: List[Journal]) -> Tuple[BytesIO, str]:
        """
        Export journals to Excel.
        
        Returns:
            (BytesIO buffer, filename)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exported Journals"
        
        # Add headers
        headers = ImportTemplate.HEADERS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add data
        row_idx = 2
        for journal in journals:
            for line in journal.lines.all():
                ws.cell(row=row_idx, column=1).value = journal.date
                ws.cell(row=row_idx, column=2).value = journal.journal_type.code
                ws.cell(row=row_idx, column=3).value = journal.reference
                ws.cell(row=row_idx, column=4).value = line.description
                ws.cell(row=row_idx, column=5).value = line.account.code
                ws.cell(row=row_idx, column=6).value = line.account.name
                ws.cell(row=row_idx, column=7).value = float(line.debit)
                ws.cell(row=row_idx, column=8).value = float(line.credit)
                row_idx += 1
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"journals_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return buffer, filename
    
    @staticmethod
    def export_journals_to_csv(journals: List[Journal]) -> Tuple[StringIO, str]:
        """
        Export journals to CSV.
        
        Returns:
            (StringIO buffer, filename)
        """
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=ImportTemplate.HEADERS)
        writer.writeheader()
        
        for journal in journals:
            for line in journal.lines.all():
                writer.writerow({
                    'Date': journal.date,
                    'Journal Type': journal.journal_type.code,
                    'Reference': journal.reference,
                    'Description': line.description,
                    'Account Code': line.account.code,
                    'Account Name': line.account.name,
                    'Debit': float(line.debit),
                    'Credit': float(line.credit),
                    'Department': '',
                    'Project': '',
                    'Cost Center': ''
                })
        
        filename = f"journals_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return output, filename
