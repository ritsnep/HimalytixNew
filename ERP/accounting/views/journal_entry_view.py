import json
import logging

from django.core.exceptions import ValidationError
from django.http import HttpRequest, JsonResponse, HttpResponse
from django.views.generic import View
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from accounting.forms.journal_line_form import JournalLineForm
from accounting.models import Journal, JournalType, JournalLine, AccountingPeriod, VoucherModeConfig
from usermanagement.utils import PermissionUtils
from accounting.forms import JournalForm, JournalLineFormSet
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from django.utils.decorators import method_decorator
from accounting.services.journal_entry_service import JournalEntryService
from accounting.validation import JournalValidationService
from utils.htmx import require_htmx
from accounting.forms.form_factory import VoucherFormFactory, get_voucher_ui_header
from accounting.services.post_journal import JournalError
from accounting.config.settings import journal_entry_settings
from accounting.models import AuditLog
from django.contrib.contenttypes.models import ContentType
from datetime import datetime

from accounting.mixins import UserOrganizationMixin
from utils.file_uploads import (
    ALLOWED_ATTACHMENT_EXTENSIONS,
    MAX_ATTACHMENT_UPLOAD_BYTES,
    iter_validated_files,
)

logger = logging.getLogger(__name__)

RECEIPT_ALLOWED_EXTENSIONS = (
    ALLOWED_ATTACHMENT_EXTENSIONS
    & {".jpg", ".jpeg", ".png", ".pdf", ".heic"}
)
if not RECEIPT_ALLOWED_EXTENSIONS:
    RECEIPT_ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}


def _prepare_receipt_file(uploaded_file, *, label: str):
    files = list(
        iter_validated_files(
            uploaded_file,
            allowed_extensions=RECEIPT_ALLOWED_EXTENSIONS,
            max_bytes=MAX_ATTACHMENT_UPLOAD_BYTES,
            allow_archive=False,
            label=label,
        )
    )
    if not files:
        raise ValidationError("Uploaded file is empty.")
    return files[0][1]

class JournalOCRView(UserOrganizationMixin, View):
    def post(self, request, *args, **kwargs):
        if not request.FILES.get('file'):
            return JsonResponse({'success': False, 'error': 'No file found.'}, status=400)

        try:
            safe_file = _prepare_receipt_file(request.FILES['file'], label="Receipt OCR")
            extracted_data = process_receipt_with_ocr(safe_file)
            return JsonResponse({'success': True, 'extracted_data': extracted_data})
        except ValidationError as exc:
            return JsonResponse({'success': False, 'error': str(exc)}, status=400)
        except Exception as e:
            logger.exception(
                "Error processing OCR for file upload: %s", e,
                extra={'user_id': request.user.pk,
                       'organization_id': self.get_organization().pk if self.get_organization() else None}
            )
            return JsonResponse({'success': False, 'error': "Failed to process receipt with OCR."}, status=500)


class JournalCheckDateView(UserOrganizationMixin, View):

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            date_str = data.get('date')
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            period = AccountingPeriod.objects.filter(
                organization=self.get_organization(),
                start_date__lte=date_obj,
                end_date__gte=date_obj,
                status='open'
            ).first()
            if not period:
                return JsonResponse({'warning': 'Date is not in an open accounting period.'})
            return JsonResponse({'success': True})
        except (json.JSONDecodeError, ValueError) as e:
            logger.error(
                "Invalid date format in JournalCheckDateView: %s. Data: %s", e, request.body,
                exc_info=True,
                extra={'user_id': request.user.pk,
                       'organization_id': self.get_organization().pk if self.get_organization() else None,
                       'request_body': request.body}
            )
            return JsonResponse({'error': 'Invalid date format'}, status=400)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred in JournalCheckDateView: %s", e,
                extra={'user_id': request.user.pk,
                       'organization_id': self.get_organization().pk if self.get_organization() else None}
            )
            return JsonResponse({'error': 'An unexpected error occurred while checking the date.'}, status=500)


class JournalHeaderFormView(UserOrganizationMixin, View):
    template_name = 'accounting/partials/journal_header_form.html'

    def get(self, request, *args, **kwargs):
        organization = self.get_organization()
        ui_schema = get_voucher_ui_header(organization)
        form = JournalForm(organization=organization, ui_schema=ui_schema)
        return render(request, self.template_name, {'form': form})


class JournalNewLineView(UserOrganizationMixin, View):
    template_name = 'accounting/partials/journal_line_form.html'

    def get(self, request, *args, **kwargs):
        organization = self.get_organization()
        # Try to fetch line UI schema if set for organization's default voucher config
        line_ui = None
        try:
            selected_config = VoucherModeConfig.objects.filter(organization=organization, is_default=True).first()
            if selected_config:
                line_ui = selected_config.resolve_ui().get('lines')
        except Exception:
            line_ui = None
        form_kwargs = {'organization': organization}
        if line_ui:
            form_kwargs['ui_schema'] = line_ui
        formset = JournalLineFormSet(queryset=JournalLine.objects.none(), prefix='lines', form_kwargs=form_kwargs)
        empty_form = formset.empty_form
        return render(request, self.template_name, {'form': empty_form})

class JournalEntryRowTemplateView(UserOrganizationMixin, View):
    """
    Returns the empty form template for a journal line.
    Used by HTMX to dynamically add new rows.
    """
    def get(self, request, *args, **kwargs):
        organization = self.get_organization()
        line_ui = None
        try:
            selected_config = VoucherModeConfig.objects.filter(organization=organization, is_default=True).first()
            if selected_config:
                line_ui = selected_config.resolve_ui().get('lines')
        except Exception:
            line_ui = None
        form_kwargs = {'organization': organization}
        if line_ui:
            form_kwargs['ui_schema'] = line_ui
        formset = JournalLineFormSet(queryset=JournalLine.objects.none(), prefix='lines', form_kwargs=form_kwargs)
        empty_form = formset.empty_form
        context = {
            'form': empty_form,
            'formset': formset, # Pass the formset to the template
            'forloop': {'counter0': '__prefix__', 'counter': '__prefix_plus_1__'} # These are placeholders for JS
        }
        return render(request, 'accounting/partials/line_empty_form.html', context)


class JournalUpdateLineView(UserOrganizationMixin, View):

    def post(self, request, *args, **kwargs):
        organization = self.get_organization()
        ui_schema = get_voucher_ui_header(organization)
        form = JournalForm(request.POST, organization=organization, ui_schema=ui_schema)
        formset = JournalLineFormSet(request.POST, queryset=JournalLine.objects.none(), prefix='lines', form_kwargs={'organization': self.get_organization()})

        if form.is_valid() and formset.is_valid():
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': 'Invalid data'})


class JournalValidateLineView(UserOrganizationMixin, View):
    """
    A view to validate a single journal line via HTMX.
    """
    def post(self, request: HttpRequest, *args, **kwargs):
        line_index = None  # Initialize line_index
        try:
            if not request.body:
                return JsonResponse({'error': 'Empty request body.'}, status=400)
            
            try:
                data = json.loads(request.body)
                line_data = data.get('line', {})
                line_index = data.get('line_index')
            except json.JSONDecodeError:
                # Fallback for non-JSON POST requests, e.g., standard form submission
                line_index = request.POST.get('line_index') # Get line_index first
                line_data = {}
                if line_index is not None:
                    for key, value in request.POST.items():
                        if key.startswith(f'lines-{line_index}-'):
                            field_name = key.split(f'lines-{line_index}-')[1]
                            line_data[field_name] = value
            
            validation_service = JournalValidationService(self.get_organization())
            # Validate a single line. The service needs to be adapted to validate a single line.
            # For now, we'll pass it as a list of one line.
            errors = validation_service.validate_journal_entry({}, [line_data])

            # Extract errors specific to this line
            line_errors = {}
            if 'lines' in errors:
                for error_info in errors['lines']:
                    if error_info.get('index') == 0: # Assuming it's the first (and only) line validated
                        line_errors = error_info['errors']
                        break

            return JsonResponse({
                'success': not bool(line_errors),
                'errors': line_errors,
                'has_errors': bool(line_errors),
                'line_index': line_index
            })
        except json.JSONDecodeError as e:
            logger.error(
                "Invalid JSON in JournalValidateLineView: %s. Request body: %s", e, request.body,
                exc_info=True,
                extra={'user_id': request.user.pk, 'organization_id': getattr(request.organization, 'pk', None), 'request_body': request.body}
            )
            return JsonResponse({'error': 'Invalid JSON data provided.'}, status=400)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred in JournalValidateLineView: %s", e,
                extra={'user_id': request.user.pk, 'organization_id': getattr(request.organization, 'pk', None)}
            )
            return JsonResponse({'error': 'An unexpected error occurred during line validation.'}, status=500)


class JournalAuditTrailView(UserOrganizationMixin, View):
    template_name = 'accounting/partials/audit_trail.html'

    def get(self, request, pk):
        journal = get_object_or_404(Journal, pk=pk, organization=self.get_organization())
        content_type = ContentType.objects.get_for_model(journal)
        audit_logs = AuditLog.objects.filter(content_type=content_type, object_id=journal.pk).order_by('-timestamp')
        context = {
            'audit_logs': audit_logs,
            'journal': journal,
        }
        return render(request, self.template_name, context)

class JournalLineDetailHXView(UserOrganizationMixin, View):
    """
    Returns the details of a single journal line for display in the side panel via HTMX.
    """
    def get(self, request, journal_id, line_index, *args, **kwargs):
        journal = get_object_or_404(Journal, pk=journal_id, organization=self.get_organization())
        line = get_object_or_404(JournalLine, journal=journal, line_number=line_index + 1) # line_number is 1-based
        
        form = JournalLineForm(instance=line, organization=self.get_organization(), prefix=f'lines-{line_index}')
        
        context = {
            'form': form,
            'line': line,
            'line_index': line_index,
        }
        return render(request, 'accounting/partials/journal_line_details.html', context)

class ValidateJournalEntryView(UserOrganizationMixin, View):
    """
    A view to validate a journal entry via HTMX.
    """

    def post(self, request: HttpRequest, *args, **kwargs):
        """
        Handles the POST request to validate the journal entry.
        """
        try:
            if not request.body:
                return JsonResponse({'error': 'Empty request body.'}, status=400)
            
            try:
                data = json.loads(request.body)
                journal_data = data.get('journal', {})
                lines_data = data.get('lines', [])
            except json.JSONDecodeError:
                journal_data = request.POST
                lines_data = []
                i = 0
                while True:
                    line_key = f'lines-{i}-account'
                    if line_key in request.POST:
                        line_data = {
                            'account': request.POST.get(f'lines-{i}-account'),
                            'description': request.POST.get(f'lines-{i}-description'),
                            'debit_amount': request.POST.get(f'lines-{i}-debit_amount'),
                            'credit_amount': request.POST.get(f'lines-{i}-credit_amount'),
                            'department': request.POST.get(f'lines-{i}-department'),
                            'project': request.POST.get(f'lines-{i}-project'),
                            'cost_center': request.POST.get(f'lines-{i}-cost_center'),
                            'exchange_rate': request.POST.get(f'lines-{i}-exchange_rate'),
                        }
                        lines_data.append(line_data)
                        i += 1
                    else:
                        break

            validation_service = JournalValidationService(self.get_organization())
            errors = validation_service.validate_journal_entry(journal_data, lines_data)

            return JsonResponse({
                'success': not errors,
                'errors': errors,
                'has_errors': bool(errors)
            })
        except json.JSONDecodeError as e:
            logger.error(
                "Invalid JSON in ValidateJournalEntryView: %s. Request body: %s", e, request.body,
               exc_info=True,
               extra={'user_id': request.user.pk, 'organization_id': getattr(request.organization, 'pk', None), 'request_body': request.body}
           )
            return JsonResponse({'error': 'Invalid JSON data provided.'}, status=400)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred in ValidateJournalEntryView: %s", e,
                extra={'user_id': request.user.pk, 'organization_id': getattr(request.organization, 'pk', None)}
            )
            return JsonResponse({'error': 'An unexpected error occurred during validation.'}, status=500)

def download_excel_template(request):
    """
    Provides a downloadable Excel template for journal entry imports
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Entry Template"

    # Define headers
    headers = [
        'Date', 'Reference', 'Description', 'Account Code', 
        'Account Name', 'Debit Amount', 'Credit Amount',
        'Department', 'Project', 'Cost Center'
    ]

    # Style the header row
    header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    header_font = Font(bold=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Add sample data row
    sample_data = [
        '2025-08-11',  # Date
        'INV-001',     # Reference
        'Sample Entry',# Description
        '1000',        # Account Code
        'Cash',        # Account Name
        '1000.00',     # Debit Amount
        '',           # Credit Amount
        'SALES',       # Department
        'PROJ-001',    # Project
        'CC-001'       # Cost Center
    ]

    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="journal_entry_template.xlsx"'

    return response

class BaseJournalEntryView(LoginRequiredMixin, View):
    template_name = 'accounting/journal_entry.html'

    def get_journal(self, pk):
        if pk:
            return get_object_or_404(Journal, pk=pk, organization=self.request.organization)
        return None

    def get_context_data(self, **kwargs):
        context = {}
        context.update(kwargs)
        context['perms'] = PermissionUtils.get_user_permissions(self.request.user, self.request.organization)
        return context

    def process_journal_submission(self, request, journal_instance=None):
        # Attach UI schema (header & lines) to the forms if available
        ui_header = get_voucher_ui_header(request.organization)
        # Determine line-level schema if present for the journal type
        line_ui = None
        try:
            journal_type = None
            if journal_instance:
                journal_type = journal_instance.journal_type
            else:
                journal_type = request.POST.get('journal_type')
            selected_config = VoucherModeConfig.objects.filter(organization=request.organization, is_default=True)
            if journal_type:
                selected_config = selected_config.filter(transaction_type=journal_type)
            selected_config = selected_config.first()
            if selected_config:
                line_ui = selected_config.resolve_ui().get('lines')
        except Exception:
            line_ui = None

        form = JournalForm(request.POST, instance=journal_instance, organization=request.organization, ui_schema=ui_header)
        formset_kwargs = {'organization': request.organization}
        if line_ui:
            formset_kwargs['ui_schema'] = line_ui
        formset = JournalLineFormSet(request.POST, instance=journal_instance, prefix='lines', form_kwargs=formset_kwargs)

        if form.is_valid() and formset.is_valid():
            validation_service = JournalValidationService(request.organization)
            journal_data = form.cleaned_data
            lines_data = formset.cleaned_data
            errors = validation_service.validate_journal_entry(journal_data, lines_data)

            if not errors:
                try:
                    with transaction.atomic():
                        journal_instance = form.save(commit=False)
                        if not journal_instance.pk:
                            journal_instance.organization = request.user.get_active_organization()
                            journal_instance.created_by = request.user
                        
                        journal_instance.save()
                        
                        formset.instance = journal_instance
                        formset.save()

                        attachments = request.FILES.getlist('attachments')
                        if attachments:
                            service = JournalEntryService(user=request.user, organization=request.organization)
                            service.add_attachments(journal_instance, attachments)
                except ValidationError as exc:
                    form.add_error(None, str(exc))
                else:
                    logger.info(
                        "Journal entry %s saved successfully by user %s",
                        journal_instance.pk, request.user.pk,
                        extra={'journal_id': journal_instance.pk, 'user_id': request.user.pk, 'organization_id': request.organization.pk}
                    )
                    return redirect('accounting:journal_detail', pk=journal_instance.pk)
            else:
                # Add errors to the form and formset
                if 'general' in errors:
                    for error in errors['general']:
                        form.add_error(None, error)
                if 'header' in errors:
                    for field, error in errors['header'].items():
                        form.add_error(field, error)
                if 'lines' in errors:
                    for error_info in errors['lines']:
                        index = error_info['index']
                        for field, error in error_info['errors'].items():
                            formset.forms[index].add_error(field, error)
                logger.warning(
                    "Journal entry validation failed for user %s, organization %s. Errors: %s",
                    request.user.pk, request.organization.pk, errors,
                    extra={'user_id': request.user.pk, 'organization_id': request.organization.pk, 'validation_errors': errors}
                )
        else:
            logger.error(
                "Journal form or formset invalid for user %s, organization %s. Form errors: %s, Formset errors: %s",
                request.user.pk, request.organization.pk, form.errors, formset.errors,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk, 'form_errors': form.errors, 'formset_errors': formset.errors}
            )
        
        context = self.get_context_data(form=form, formset=formset, journal=journal_instance)
        return render(request, self.template_name, context)

class JournalEntryDetailView(BaseJournalEntryView):
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        if not pk:
            # If no PK is provided, redirect to the create view
            return redirect('accounting:journal_entry_new')
        
        journal = self.get_journal(pk)
        ui_header = get_voucher_ui_header(request.organization, journal_type=journal.journal_type if journal else None)
        form = JournalForm(instance=journal, organization=request.organization, ui_schema=ui_header)
        if journal:
            # Attempt to get line UI schema for this journal type
            line_ui = None
            try:
                selected_config = VoucherModeConfig.objects.filter(organization=request.organization, is_default=True, transaction_type=journal.journal_type).first()
                if selected_config:
                    line_ui = selected_config.resolve_ui().get('lines')
            except Exception:
                line_ui = None
            form_kwargs = {'organization': request.organization}
            if line_ui:
                form_kwargs['ui_schema'] = line_ui
            formset = JournalLineFormSet(queryset=journal.lines.all(), prefix='lines', form_kwargs=form_kwargs)
        else:
            formset = JournalLineFormSet(queryset=JournalLine.objects.none(), prefix='lines')
        context = self.get_context_data(form=form, formset=formset, journal=journal)
        return render(request, self.template_name, context)

class JournalEntryCreateView(BaseJournalEntryView):
    def get(self, request, *args, **kwargs):
        ui_header = get_voucher_ui_header(request.organization)
        form = JournalForm(organization=request.organization, ui_schema=ui_header)
        # Get default line UI schema
        line_ui = None
        try:
            selected_config = VoucherModeConfig.objects.filter(organization=request.organization, is_default=True).first()
            if selected_config:
                line_ui = selected_config.resolve_ui().get('lines')
        except Exception:
            line_ui = None
        form_kwargs = {'organization': request.organization}
        if line_ui:
            form_kwargs['ui_schema'] = line_ui
        formset = JournalLineFormSet(queryset=JournalLine.objects.none(), prefix='lines', form_kwargs=form_kwargs)
        context = self.get_context_data(form=form, formset=formset)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        return self.process_journal_submission(request)

class JournalEntryUpdateView(BaseJournalEntryView):
    def get(self, request, *args, **kwargs):
        journal = self.get_journal(kwargs.get('pk'))
        ui_header = get_voucher_ui_header(request.organization, journal_type=journal.journal_type if journal else None)
        form = JournalForm(instance=journal, organization=request.organization, ui_schema=ui_header)
        # Attempt to get line UI schema for journal type and pass to formset
        line_ui = None
        try:
            selected_config = VoucherModeConfig.objects.filter(organization=request.organization, is_default=True, transaction_type=journal.journal_type).first()
            if selected_config:
                line_ui = selected_config.resolve_ui().get('lines')
        except Exception:
            line_ui = None
        form_kwargs = {'organization': request.organization}
        if line_ui:
            form_kwargs['ui_schema'] = line_ui
        formset = JournalLineFormSet(queryset=journal.lines.all(), prefix='lines', form_kwargs=form_kwargs)
        context = self.get_context_data(form=form, formset=formset, journal=journal)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        journal = self.get_journal(kwargs.get('pk'))
        return self.process_journal_submission(request, journal_instance=journal)

class JournalConfigChangeView(LoginRequiredMixin, View):
    """
    Handles AJAX requests for journal configuration changes
    """
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            config_type = data.get('config_type')
            config_id = data.get('config_id')
            
            if config_type == 'journal_type':
                journal_type = JournalType.objects.get(
                    pk=config_id,
                    organization=request.organization
                )
                
                return JsonResponse({
                    'success': True,
                    'data': {
                        'auto_numbering': journal_type.auto_numbering,
                        'prefix': journal_type.prefix,
                        'suffix': journal_type.suffix,
                        'next_number': journal_type.next_number,
                        'currency_enabled': journal_type.currency_enabled,
                        'default_currency': journal_type.default_currency,
                        'exchange_rate_required': journal_type.exchange_rate_required
                    }
                })
            
            return JsonResponse({
                'success': False,
                'message': 'Invalid configuration type'
            }, status=400)
            
        except JournalType.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Journal type not found'
            }, status=404)
        except json.JSONDecodeError as e:
            logger.error(
                "Invalid JSON in JournalConfigChangeView: %s. Request body: %s", e, request.body,
                exc_info=True,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk, 'request_body': request.body}
            )
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON data provided.'
            }, status=400)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred in JournalConfigChangeView: %s", e,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk}
            )
            return JsonResponse({
                'success': False,
                'message': 'An unexpected error occurred while changing configuration.'
            }, status=500)

from .views_mixins import UserOrganizationMixin


class JournalLineAddView(UserOrganizationMixin, View):
    """
    Handles adding new lines to a journal entry via HTMX
    """
    template_name = 'accounting/partials/line_empty_form.html'

    @method_decorator(require_htmx)
    def post(self, request, *args, **kwargs):
        try:
            # Get the current number of lines
            total_forms = int(request.POST.get('lines-TOTAL_FORMS', 0))

            # Check if the maximum number of lines has been reached
            if total_forms >= journal_entry_settings.max_lines_per_entry:
                return HttpResponse(
                    '<div class="alert alert-danger">Maximum number of lines reached.</div>',
                    status=400
                )

            # Create a new empty form with the next index
            formset = JournalLineFormSet(
                queryset=JournalLine.objects.none(),
                prefix='lines',
                form_kwargs={'organization': self.get_organization()}
            )
            empty_form = formset.empty_form

            context = {
                'form': empty_form,
                'forloop': {'counter0': total_forms, 'counter': total_forms + 1}
            }

            return render(request, self.template_name, context)

        except Exception as e:
            logger.exception(
                "An unexpected error occurred in JournalLineAddView: %s", e,
                extra={'user_id': request.user.pk,
                       'organization_id': self.get_organization().pk if self.get_organization() else None}
            )
            return JsonResponse({
                'success': False,
                'message': 'An unexpected error occurred while adding a new line.'
            }, status=500)


class UploadReceiptView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        if not request.FILES.get('receipt'):
            return JsonResponse({'success': False, 'error': 'No receipt file found.'}, status=400)

        try:
            safe_file = _prepare_receipt_file(request.FILES['receipt'], label="Receipt upload")
            extracted_data = process_receipt_with_ocr(safe_file)
            return JsonResponse({'success': True, 'extracted_data': extracted_data})
        except ValidationError as exc:
            return JsonResponse({'success': False, 'error': str(exc)}, status=400)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred in UploadReceiptView: %s", e,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk}
            )
            return JsonResponse({'success': False, 'error': 'An unexpected error occurred during receipt upload.'}, status=500)

class JournalSaveView(LoginRequiredMixin, View):
    """Handles saving journal entries, including validation and line items."""
    
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            journal_data = data.get('journal', {})
            lines_data = data.get('lines', [])
            journal_id = journal_data.get('id')

            validation_service = JournalValidationService(request.organization)
            errors = validation_service.validate_journal_entry(journal_data, lines_data)

            if errors:
                return JsonResponse({
                    'success': False,
                    'errors': errors
                }, status=400)

            service = JournalEntryService(user=request.user, organization=request.organization)

            if journal_id:
                journal = get_object_or_404(Journal, pk=journal_id, organization=request.organization)
                attachments = request.FILES.getlist('attachments')
                journal = service.update_journal_entry(journal, journal_data, lines_data, attachments)
            else:
                attachments = request.FILES.getlist('attachments')
                journal = service.create_journal_entry(journal_data, lines_data, attachments)

            return JsonResponse({
                'success': True,
                'journal_id': journal.pk,
                'message': 'Journal saved successfully'
            })

        except (ValueError, JournalError, ValidationError) as e:
            logger.error(
                "Journal save failed for user %s, organization %s: %s",
                request.user.pk, request.organization.pk, e,
                exc_info=True,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk, 'error_message': str(e)}
            )
            return JsonResponse({
                'success': False,
                'errors': {'general': str(e)}
            }, status=400)
        except Journal.DoesNotExist:
            logger.warning(
                "Attempted to save non-existent journal_id=%s for user %s, organization %s",
                journal_id, request.user.pk, request.organization.pk,
                extra={'journal_id': journal_id, 'user_id': request.user.pk, 'organization_id': request.organization.pk}
            )
            return JsonResponse({
                'success': False,
                'errors': {'journal': 'Journal not found'}
            }, status=404)
        except Exception as e:
            logger.exception(
                "An unexpected error occurred during journal save for user %s, organization %s: %s",
                request.user.pk, request.organization.pk, e,
                extra={'user_id': request.user.pk, 'organization_id': request.organization.pk}
            )
            return JsonResponse({
                'success': False,
                'errors': {'general': 'An unexpected error occurred while saving the journal.'}
            }, status=500)

class JournalEntryValidateView(UserOrganizationMixin, View):
    """
    HTMX endpoint to trigger background validation for a journal entry.
    Returns the updated validation panel partial.
    """
    def post(self, request, *args, **kwargs):
        journal_id = request.POST.get('journal_id')
        journal = None
        if journal_id:
            journal = get_object_or_404(Journal, pk=journal_id, organization=self.get_organization())

        form = JournalForm(request.POST, instance=journal, organization=self.get_organization())
        formset = JournalLineFormSet(request.POST, prefix='lines', form_kwargs={'organization': self.get_organization()})

        # Ensure forms are valid to populate cleaned_data
        form_is_valid = form.is_valid()
        formset_is_valid = formset.is_valid()

        validation_service = JournalValidationService(self.get_organization())
        lines_cleaned_data = formset.cleaned_data if formset_is_valid else []
        validation_results = validation_service.validate_journal_entry(form.cleaned_data, lines_cleaned_data)

        context = {
            'validation_results': validation_results,
            'journal': journal, # Pass journal for potential context in the partial
        }
        return render(request, 'accounting/partials/_journal_validation_panel.html', context)