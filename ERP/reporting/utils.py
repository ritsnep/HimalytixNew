from __future__ import annotations

import csv
import io
import re
from typing import Any, Dict, Iterable, List, Tuple

from django.template import engines
from django.template.loader import render_to_string
from django.utils import timezone


SCRIPT_TAG_RE = re.compile(r"<script.*?>.*?</script>", re.IGNORECASE | re.DOTALL)
JS_PROTO_RE = re.compile(r"javascript:", re.IGNORECASE)


def sanitize_template_html(html: str) -> str:
    """Remove script tags and obvious JavaScript protocol usage."""
    if not html:
        return ""
    cleaned = SCRIPT_TAG_RE.sub("", html)
    cleaned = JS_PROTO_RE.sub("", cleaned)
    return cleaned


def render_template_string(engine_name: str, template_html: str, context: Dict[str, Any], request=None) -> str:
    """Render a template string using the configured engine."""
    if engine_name == "jinja":
        try:
            jinja_engine = engines["jinja2"]
        except Exception as exc:  # noqa: BLE001
            raise ValueError("Jinja2 engine is not configured. Update TEMPLATES to include jinja2.") from exc
        template = jinja_engine.from_string(template_html)
        return template.render(context)

    django_engine = engines["django"]
    template = django_engine.from_string(template_html)
    return template.render(context, request)


def render_base_template(template_name: str, context: Dict[str, Any], request=None) -> str:
    """Render a filesystem template (base fallback)."""
    return render_to_string(template_name, context=context, request=request)


def export_csv(columns: Iterable[str], rows: Iterable[Iterable[Any]], title: str = "report") -> Tuple[io.BytesIO, str]:
    """Export tabular data to CSV."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    if title:
        writer.writerow([title])
        writer.writerow([f"Generated: {timezone.now()}"])
        writer.writerow([])
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    data = io.BytesIO(buffer.getvalue().encode("utf-8"))
    filename = f"{title.replace(' ', '_').lower()}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return data, filename


def export_excel(columns: List[str], rows: List[List[Any]], title: str = "report"):
    """Export tabular data to Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # noqa: BLE001
        raise ImportError("openpyxl is required for Excel export.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append([title])
    ws.append([f"Generated: {timezone.now()}"])
    ws.append([])

    header_font = Font(bold=True)
    ws.append(columns)
    for col, _ in enumerate(columns, 1):
        ws.cell(row=ws.max_row, column=col).font = header_font

    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{title.replace(' ', '_').lower()}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def export_pdf(html: str, title: str = "report"):
    """Export rendered HTML to PDF using WeasyPrint."""
    try:
        from weasyprint import HTML
    except ImportError as exc:  # noqa: BLE001
        raise ImportError("WeasyPrint is required for PDF export.") from exc

    output = io.BytesIO()
    HTML(string=html).write_pdf(output)
    output.seek(0)
    filename = f"{title.replace(' ', '_').lower()}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return output, filename
