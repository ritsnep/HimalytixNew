"""
Data Import/Export Utilities

Provides comprehensive data import/export functionality for the accounting system,
supporting CSV, Excel, and JSON formats with validation and error handling.
"""

import csv
import json
import logging
from io import StringIO, BytesIO
from typing import Optional, Dict, List, Any, Tuple, Union, BinaryIO, TextIO
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

import openpyxl
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone

from .organization import OrganizationService
from .validation import BusinessValidator

logger = logging.getLogger(__name__)


class DataImporter:
    """
    Generic data import utility with validation and error handling.
    """

    def __init__(self, organization: Any, user: Any):
        self.organization = organization
        self.user = user
        self.errors = []
        self.warnings = []
        self.processed_count = 0

    def import_accounts(self, file_obj: BinaryIO, file_format: str = 'auto') -> Dict[str, Any]:
        """
        Import chart of accounts from file.

        Args:
            file_obj: File-like object containing the data
            file_format: Format of the file ('csv', 'excel', 'auto')

        Returns:
            Dictionary with import results

        Usage:
            # Import accounts from uploaded file
            with open('accounts.csv', 'rb') as f:
                result = importer.import_accounts(f, 'csv')
        """
        if file_format == 'auto':
            file_format = self._detect_format(file_obj)

        data = self._parse_file(file_obj, file_format)
        return self._import_accounts_data(data)

    def import_transactions(self, file_obj: BinaryIO, file_format: str = 'auto') -> Dict[str, Any]:
        """
        Import transactions/journals from file.

        Args:
            file_obj: File-like object containing transaction data
            file_format: Format of the file

        Returns:
            Dictionary with import results
        """
        if file_format == 'auto':
            file_format = self._detect_format(file_obj)

        data = self._parse_file(file_obj, file_format)
        return self._import_transactions_data(data)

    def import_customers(self, file_obj: BinaryIO, file_format: str = 'auto') -> Dict[str, Any]:
        """
        Import customers from file.

        Args:
            file_obj: File-like object containing customer data
            file_format: Format of the file

        Returns:
            Dictionary with import results
        """
        if file_format == 'auto':
            file_format = self._detect_format(file_obj)

        data = self._parse_file(file_obj, file_format)
        return self._import_customers_data(data)

    def _detect_format(self, file_obj: BinaryIO) -> str:
        """Detect file format from content."""
        # Read first few bytes to detect format
        pos = file_obj.tell()
        sample = file_obj.read(1024).decode('utf-8', errors='ignore')
        file_obj.seek(pos)

        if sample.startswith('\ufeff'):  # BOM for UTF-8 Excel
            return 'excel'
        elif ',' in sample and '\n' in sample:
            return 'csv'
        elif sample.strip().startswith(('{', '[')):
            return 'json'
        else:
            # Default to CSV
            return 'csv'

    def _parse_file(self, file_obj: BinaryIO, file_format: str) -> List[Dict[str, Any]]:
        """Parse file content into list of dictionaries."""
        if file_format == 'csv':
            return self._parse_csv(file_obj)
        elif file_format == 'excel':
            return self._parse_excel(file_obj)
        elif file_format == 'json':
            return self._parse_json(file_obj)
        else:
            raise ValueError(f"Unsupported file format: {file_format}")

    def _parse_csv(self, file_obj: BinaryIO) -> List[Dict[str, Any]]:
        """Parse CSV file."""
        content = file_obj.read().decode('utf-8')
        reader = csv.DictReader(StringIO(content))
        return list(reader)

    def _parse_excel(self, file_obj: BinaryIO) -> List[Dict[str, Any]]:
        """Parse Excel file."""
        workbook = openpyxl.load_workbook(file_obj, read_only=True)
        sheet = workbook.active

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value or '')

        # Parse data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value
            data.append(row_data)

        return data

    def _parse_json(self, file_obj: BinaryIO) -> List[Dict[str, Any]]:
        """Parse JSON file."""
        content = file_obj.read().decode('utf-8')
        return json.loads(content)

    def _import_accounts_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Import accounts data with validation."""
        from .coa import COAService

        results = {
            'success': True,
            'created': 0,
            'updated': 0,
            'errors': [],
            'warnings': []
        }

        for row_num, row in enumerate(data, 1):
            try:
                with transaction.atomic():
                    account = self._process_account_row(row)
                    if account:
                        if account.pk:  # Existing account updated
                            results['updated'] += 1
                        else:  # New account created
                            results['created'] += 1

            except Exception as e:
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'data': row
                })
                logger.error(f"Error importing account at row {row_num}: {e}")

        results['success'] = len(results['errors']) == 0
        return results

    def _process_account_row(self, row: Dict[str, Any]) -> Optional[Any]:
        """Process a single account import row."""
        from accounting.models import AccountType

        # Extract required fields
        account_code = row.get('account_code', '').strip()
        account_name = row.get('account_name', '').strip()
        account_type_name = row.get('account_type', '').strip()

        if not account_code or not account_name or not account_type_name:
            raise ValueError("Account code, name, and type are required")

        # Find account type
        try:
            account_type = AccountType.objects.get(name__iexact=account_type_name)
        except AccountType.DoesNotExist:
            raise ValueError(f"Account type '{account_type_name}' not found")

        # Check if account already exists
        from accounting.models import ChartOfAccount
        existing_account = ChartOfAccount.objects.filter(
            organization=self.organization,
            account_code=account_code
        ).first()

        if existing_account:
            # Update existing account
            existing_account.account_name = account_name
            existing_account.save()
            return existing_account
        else:
            # Create new account
            return COAService.create_account(
                organization=self.organization,
                account_type=account_type,
                name=account_name,
                description=row.get('description', ''),
                is_active=row.get('is_active', 'true').lower() == 'true'
            )

    def _import_transactions_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Import transactions data."""
        # Implementation for transaction import
        # This would be more complex as it involves creating journals
        return {
            'success': False,
            'message': 'Transaction import not yet implemented'
        }

    def _import_customers_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Import customers data."""
        results = {
            'success': True,
            'created': 0,
            'updated': 0,
            'errors': [],
            'warnings': []
        }

        for row_num, row in enumerate(data, 1):
            try:
                with transaction.atomic():
                    customer = self._process_customer_row(row)
                    if customer:
                        if hasattr(customer, 'pk') and customer.pk:  # Existing customer updated
                            results['updated'] += 1
                        else:  # New customer created
                            results['created'] += 1

            except Exception as e:
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'data': row
                })
                logger.error(f"Error importing customer at row {row_num}: {e}")

        results['success'] = len(results['errors']) == 0
        return results

    def _process_customer_row(self, row: Dict[str, Any]) -> Optional[Any]:
        """Process a single customer import row."""
        from accounting.models import Customer

        # Extract required fields
        code = row.get('code', '').strip()
        display_name = row.get('display_name', '').strip()

        if not code or not display_name:
            raise ValueError("Customer code and display name are required")

        # Check if customer already exists
        existing_customer = Customer.objects.filter(
            organization=self.organization,
            code=code
        ).first()

        if existing_customer:
            # Update existing customer
            existing_customer.display_name = display_name
            existing_customer.save()
            return existing_customer
        else:
            # Create new customer
            return Customer.objects.create(
                organization=self.organization,
                code=code,
                display_name=display_name,
                created_by=self.user
            )

    def get_import_summary(self) -> Dict[str, Any]:
        """Get summary of import operation."""
        return {
            'processed_count': self.processed_count,
            'errors': self.errors,
            'warnings': self.warnings,
            'success_count': self.processed_count - len(self.errors)
        }


class DataExporter:
    """
    Generic data export utility supporting multiple formats.
    """

    def __init__(self, organization: Any = None):
        self.organization = organization

    def export_accounts(self, queryset=None, file_format: str = 'csv') -> Union[str, bytes]:
        """
        Export chart of accounts to specified format.

        Args:
            queryset: QuerySet of accounts to export (optional)
            file_format: Export format ('csv', 'excel', 'json')

        Returns:
            File content as string or bytes

        Usage:
            # Export all accounts as CSV
            csv_content = exporter.export_accounts(file_format='csv')
        """
        if queryset is None:
            from accounting.models import ChartOfAccount
            queryset = ChartOfAccount.active_accounts.for_org(self.organization)

        data = self._prepare_accounts_data(queryset)

        if file_format == 'csv':
            return self._export_csv(data, self._get_accounts_headers())
        elif file_format == 'excel':
            return self._export_excel(data, self._get_accounts_headers(), 'Accounts')
        elif file_format == 'json':
            return json.dumps(data, indent=2, default=str)
        else:
            raise ValueError(f"Unsupported export format: {file_format}")

    def export_transactions(self, queryset=None, file_format: str = 'csv') -> Union[str, bytes]:
        """
        Export transactions to specified format.

        Args:
            queryset: QuerySet of transactions to export
            file_format: Export format

        Returns:
            File content
        """
        if queryset is None:
            from accounting.models import Journal
            queryset = Journal.objects.filter(organization=self.organization)

        data = self._prepare_transactions_data(queryset)

        if file_format == 'csv':
            return self._export_csv(data, self._get_transactions_headers())
        elif file_format == 'excel':
            return self._export_excel(data, self._get_transactions_headers(), 'Transactions')
        elif file_format == 'json':
            return json.dumps(data, indent=2, default=str)
        else:
            raise ValueError(f"Unsupported export format: {file_format}")

    def _prepare_accounts_data(self, queryset) -> List[Dict[str, Any]]:
        """Prepare accounts data for export."""
        data = []
        for account in queryset:
            data.append({
                'account_code': account.account_code,
                'account_name': account.account_name,
                'account_type': account.account_type.name if account.account_type else '',
                'nature': account.account_type.nature if account.account_type else '',
                'current_balance': str(account.current_balance or 0),
                'is_active': 'Yes' if account.is_active else 'No',
                'description': account.description or '',
            })
        return data

    def _prepare_transactions_data(self, queryset) -> List[Dict[str, Any]]:
        """Prepare transactions data for export."""
        data = []
        for journal in queryset:
            for line in journal.lines.all():
                data.append({
                    'journal_number': journal.journal_number,
                    'journal_date': journal.journal_date.isoformat(),
                    'account_code': line.account.account_code,
                    'account_name': line.account.account_name,
                    'description': line.description or '',
                    'debit_amount': str(line.debit_amount or 0),
                    'credit_amount': str(line.credit_amount or 0),
                    'currency': journal.currency_code,
                    'status': journal.status,
                })
        return data

    def _get_accounts_headers(self) -> List[str]:
        """Get headers for accounts export."""
        return [
            'account_code', 'account_name', 'account_type', 'nature',
            'current_balance', 'is_active', 'description'
        ]

    def _get_transactions_headers(self) -> List[str]:
        """Get headers for transactions export."""
        return [
            'journal_number', 'journal_date', 'account_code', 'account_name',
            'description', 'debit_amount', 'credit_amount', 'currency', 'status'
        ]

    def _export_csv(self, data: List[Dict[str, Any]], headers: List[str]) -> str:
        """Export data as CSV string."""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()

    def _export_excel(self, data: List[Dict[str, Any]], headers: List[str], sheet_name: str) -> bytes:
        """Export data as Excel bytes."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # Write headers
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                sheet.cell(row=row_num, column=col_num, value=value)

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class BulkOperations:
    """
    Utilities for bulk data operations.
    """

    @staticmethod
    def bulk_update_account_status(organization: Any, account_ids: List[int], is_active: bool) -> int:
        """
        Bulk update account active status.

        Args:
            organization: Organization instance
            account_ids: List of account IDs to update
            is_active: New active status

        Returns:
            Number of accounts updated
        """
        from accounting.models import ChartOfAccount

        queryset = ChartOfAccount.objects.filter(
            organization=organization,
            pk__in=account_ids
        )

        return queryset.update(is_active=is_active, updated_at=timezone.now())

    @staticmethod
    def bulk_archive_old_records(organization: Any, model_class: Any, days_old: int = 365) -> int:
        """
        Bulk archive old records.

        Args:
            organization: Organization instance
            model_class: Django model class
            days_old: Records older than this many days

        Returns:
            Number of records archived
        """
        cutoff_date = timezone.now() - timezone.timedelta(days=days_old)

        queryset = model_class.objects.filter(
            organization=organization,
            created_at__lt=cutoff_date
        )

        if hasattr(model_class, 'is_archived'):
            return queryset.update(is_archived=True, archived_at=timezone.now())

        return 0

    @staticmethod
    def bulk_validate_data(queryset, validation_func) -> Tuple[int, List[Dict[str, Any]]]:
        """
        Bulk validate data using a validation function.

        Args:
            queryset: QuerySet to validate
            validation_func: Function that takes a model instance and returns (is_valid, error)

        Returns:
            Tuple of (valid_count, errors_list)
        """
        valid_count = 0
        errors = []

        for instance in queryset:
            is_valid, error_msg = validation_func(instance)
            if is_valid:
                valid_count += 1
            else:
                errors.append({
                    'id': instance.pk,
                    'error': error_msg,
                    'data': str(instance)
                })

        return valid_count, errors


# Import template generators
def generate_accounts_import_template() -> str:
    """
    Generate CSV template for accounts import.

    Returns:
        CSV template content
    """
    headers = ['account_code', 'account_name', 'account_type', 'description', 'is_active']
    sample_data = [
        ['1000', 'Cash', 'Asset', 'Petty cash account', 'true'],
        ['2000', 'Accounts Payable', 'Liability', 'Vendor payable account', 'true'],
        ['3000', 'Capital', 'Equity', 'Owner equity account', 'true'],
        ['4000', 'Sales Revenue', 'Income', 'Main sales account', 'true'],
        ['5000', 'Office Expenses', 'Expense', 'General office expenses', 'true'],
    ]

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(sample_data)

    return output.getvalue()


def generate_customers_import_template() -> str:
    """
    Generate CSV template for customers import.

    Returns:
        CSV template content
    """
    headers = ['code', 'display_name', 'tax_id', 'email', 'phone_number']
    sample_data = [
        ['CUST001', 'ABC Corporation', '123456789', 'contact@abc.com', '+1-555-0123'],
        ['CUST002', 'XYZ Ltd', '987654321', 'info@xyz.com', '+1-555-0456'],
    ]

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(sample_data)

    return output.getvalue()
