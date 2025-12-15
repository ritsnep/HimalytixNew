"""
Report Generation Utilities

Provides comprehensive reporting functionality for financial statements,
custom reports, and data export capabilities for the accounting system.
"""

from typing import Optional, Dict, List, Any, Tuple, Union, BinaryIO
from decimal import Decimal
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
from django.db.models import Sum, Q, F, Case, When, Value, Count, Avg
from django.db.models.functions import TruncMonth, TruncQuarter, TruncYear
from django.utils import timezone
from django.template.loader import render_to_string

from .organization import OrganizationService
from .financial import BalanceCalculator, TaxCalculator, FinancialRatioCalculator
from .cache_utils import cached_organization_data, CacheManager


class ReportGenerator:
    """
    Comprehensive report generation service supporting various financial reports,
    custom queries, and multiple export formats.
    """

    def __init__(self, organization: Any):
        self.organization = organization

    @cached_organization_data(timeout=1800)  # Cache for 30 minutes
    def generate_trial_balance(
        self,
        as_of_date: Optional[date] = None,
        include_children: bool = True,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate trial balance report.

        Args:
            as_of_date: Date for balance calculation
            include_children: Include child account balances
            format_type: 'data', 'excel', 'pdf', 'html'

        Returns:
            Report data or file content
        """
        trial_balance = BalanceCalculator.calculate_trial_balance(
            self.organization, as_of_date, include_children
        )

        if format_type == 'data':
            return trial_balance
        elif format_type == 'excel':
            return self._export_trial_balance_excel(trial_balance, as_of_date)
        elif format_type == 'html':
            return self._render_trial_balance_html(trial_balance, as_of_date)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    @cached_organization_data(timeout=1800)
    def generate_balance_sheet(
        self,
        as_of_date: Optional[date] = None,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate balance sheet report.

        Args:
            as_of_date: Date for report
            format_type: Output format

        Returns:
            Report data or file content
        """
        balance_sheet = BalanceCalculator.calculate_balance_sheet(
            self.organization, as_of_date
        )

        if format_type == 'data':
            return balance_sheet
        elif format_type == 'excel':
            return self._export_balance_sheet_excel(balance_sheet, as_of_date)
        elif format_type == 'html':
            return self._render_balance_sheet_html(balance_sheet, as_of_date)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    @cached_organization_data(timeout=1800)
    def generate_income_statement(
        self,
        start_date: date,
        end_date: date,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate income statement for period.

        Args:
            start_date: Period start date
            end_date: Period end date
            format_type: Output format

        Returns:
            Report data or file content
        """
        income_statement = BalanceCalculator.calculate_income_statement(
            self.organization, start_date, end_date
        )

        if format_type == 'data':
            return income_statement
        elif format_type == 'excel':
            return self._export_income_statement_excel(income_statement)
        elif format_type == 'html':
            return self._render_income_statement_html(income_statement)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_cash_flow_statement(
        self,
        start_date: date,
        end_date: date,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate cash flow statement.

        Args:
            start_date: Period start date
            end_date: Period end date
            format_type: Output format

        Returns:
            Report data or file content
        """
        cash_flow_data = self._calculate_cash_flow(start_date, end_date)

        if format_type == 'data':
            return cash_flow_data
        elif format_type == 'excel':
            return self._export_cash_flow_excel(cash_flow_data)
        elif format_type == 'html':
            return self._render_cash_flow_html(cash_flow_data)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_account_ledger(
        self,
        account_id: int,
        start_date: date,
        end_date: date,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate account ledger report.

        Args:
            account_id: Account ID
            start_date: Period start date
            end_date: Period end date
            format_type: Output format

        Returns:
            Report data or file content
        """
        from accounting.models import ChartOfAccount
        from .coa import COAService

        try:
            account = ChartOfAccount.objects.get(
                pk=account_id,
                organization=self.organization
            )
        except ChartOfAccount.DoesNotExist:
            raise ValueError(f"Account {account_id} not found")

        # Get opening balance
        opening_balance = COAService.get_account_balance(account, start_date - timedelta(days=1))

        # Get transactions
        transactions = COAService.get_account_transactions(
            account, start_date, end_date
        ).order_by('transaction_date')

        # Build ledger entries
        ledger_entries = []
        running_balance = opening_balance

        for transaction in transactions:
            entry = {
                'date': transaction.transaction_date,
                'description': transaction.description or '',
                'debit': transaction.debit_amount or Decimal('0'),
                'credit': transaction.credit_amount or Decimal('0'),
                'balance': running_balance,
            }

            if transaction.debit_amount:
                running_balance += transaction.debit_amount
            if transaction.credit_amount:
                running_balance -= transaction.credit_amount

            entry['running_balance'] = running_balance
            ledger_entries.append(entry)

        ledger_data = {
            'account': {
                'code': account.account_code,
                'name': account.account_name,
                'type': account.account_type.name if account.account_type else '',
            },
            'period': {'start': start_date, 'end': end_date},
            'opening_balance': opening_balance,
            'closing_balance': running_balance,
            'entries': ledger_entries,
        }

        if format_type == 'data':
            return ledger_data
        elif format_type == 'excel':
            return self._export_ledger_excel(ledger_data)
        elif format_type == 'html':
            return self._render_ledger_html(ledger_data)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_tax_report(
        self,
        start_date: date,
        end_date: date,
        tax_type: Optional[str] = None,
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate tax report.

        Args:
            start_date: Period start date
            end_date: Period end date
            tax_type: Specific tax type to report on
            format_type: Output format

        Returns:
            Report data or file content
        """
        tax_data = self._calculate_tax_summary(start_date, end_date, tax_type)

        if format_type == 'data':
            return tax_data
        elif format_type == 'excel':
            return self._export_tax_report_excel(tax_data)
        elif format_type == 'html':
            return self._render_tax_report_html(tax_data)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_custom_report(
        self,
        query_config: Dict[str, Any],
        format_type: str = 'data'
    ) -> Union[Dict[str, Any], bytes]:
        """
        Generate custom report based on configuration.

        Args:
            query_config: Report configuration dictionary
            format_type: Output format

        Returns:
            Report data or file content
        """
        # Parse query configuration
        model_name = query_config.get('model')
        filters = query_config.get('filters', {})
        aggregations = query_config.get('aggregations', {})
        group_by = query_config.get('group_by', [])
        date_field = query_config.get('date_field', 'created_at')
        date_range = query_config.get('date_range')

        # Build queryset
        queryset = self._build_custom_queryset(
            model_name, filters, date_range, date_field
        )

        # Apply aggregations and grouping
        if aggregations:
            result = queryset.aggregate(**aggregations)
        elif group_by:
            result = self._apply_group_by(queryset, group_by, aggregations)
        else:
            result = list(queryset.values())

        report_data = {
            'config': query_config,
            'data': result,
            'generated_at': timezone.now(),
        }

        if format_type == 'data':
            return report_data
        elif format_type == 'excel':
            return self._export_custom_report_excel(report_data)
        else:
            raise ValueError(f"Unsupported format for custom reports: {format_type}")

    def generate_financial_ratios(
        self,
        as_of_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Generate financial ratios report.

        Args:
            as_of_date: Date for ratio calculation

        Returns:
            Financial ratios data
        """
        # Get balance sheet data
        balance_sheet = BalanceCalculator.calculate_balance_sheet(
            self.organization, as_of_date
        )

        # Get income statement data (last 12 months)
        end_date = as_of_date or timezone.localdate()
        start_date = end_date - timedelta(days=365)

        income_statement = BalanceCalculator.calculate_income_statement(
            self.organization, start_date, end_date
        )

        # Extract key figures
        current_assets = balance_sheet['assets']['total']
        current_liabilities = balance_sheet['liabilities_and_equity']['liabilities']['total']
        total_equity = balance_sheet['liabilities_and_equity']['equity']['total']
        net_income = income_statement.get('net_income', Decimal('0'))
        sales = Decimal('0')  # Would need to calculate from revenue accounts

        # Calculate ratios
        ratios = {
            'liquidity_ratios': FinancialRatioCalculator.calculate_liquidity_ratios(
                current_assets, current_liabilities, current_assets * Decimal('0.3')  # Estimate cash
            ),
            'profitability_ratios': FinancialRatioCalculator.calculate_profitability_ratios(
                net_income, current_assets + balance_sheet['liabilities_and_equity']['total'],
                total_equity, sales
            ),
        }

        return {
            'as_of_date': end_date,
            'ratios': ratios,
            'key_figures': {
                'current_assets': current_assets,
                'current_liabilities': current_liabilities,
                'total_equity': total_equity,
                'net_income': net_income,
            }
        }

    def _calculate_cash_flow(self, start_date: date, end_date: date) -> Dict[str, Any]:
        """Calculate cash flow statement data."""
        # This is a simplified implementation
        # In practice, this would involve complex cash flow calculations

        # Get net income
        income_statement = BalanceCalculator.calculate_income_statement(
            self.organization, start_date, end_date
        )
        net_income = income_statement.get('net_income', Decimal('0'))

        # Simplified cash flow categories
        cash_flow_data = {
            'period': {'start': start_date, 'end': end_date},
            'operating_activities': {
                'net_income': net_income,
                'adjustments': [],  # Would calculate non-cash items
                'net_cash_operating': net_income,
            },
            'investing_activities': {
                'items': [],  # Asset purchases, sales, etc.
                'net_cash_investing': Decimal('0'),
            },
            'financing_activities': {
                'items': [],  # Loans, equity changes, etc.
                'net_cash_financing': Decimal('0'),
            },
            'net_cash_change': net_income,
        }

        return cash_flow_data

    def _calculate_tax_summary(self, start_date: date, end_date: date, tax_type: Optional[str] = None) -> Dict[str, Any]:
        """Calculate tax summary for period."""
        # This would query tax-related transactions
        # Simplified implementation
        return {
            'period': {'start': start_date, 'end': end_date},
            'tax_type': tax_type,
            'total_tax_collected': Decimal('0'),
            'total_tax_paid': Decimal('0'),
            'net_tax_position': Decimal('0'),
            'tax_details': [],
        }

    def _build_custom_queryset(self, model_name: str, filters: Dict[str, Any],
                             date_range: Optional[Dict[str, Any]], date_field: str = 'created_at'):
        """Build queryset for custom reports."""
        # Import model dynamically
        from django.apps import apps
        try:
            model = apps.get_model('accounting', model_name)
        except:
            raise ValueError(f"Model {model_name} not found")

        queryset = model.objects.all()

        # Apply organization filter
        queryset = OrganizationService.filter_queryset_by_org(queryset, self.organization)

        # Apply custom filters
        for field, value in filters.items():
            if isinstance(value, dict):
                # Range or complex filter
                if 'gte' in value:
                    queryset = queryset.filter(**{f"{field}__gte": value['gte']})
                if 'lte' in value:
                    queryset = queryset.filter(**{f"{field}__lte": value['lte']})
            else:
                queryset = queryset.filter(**{field: value})

        # Apply date range
        if date_range:
            if 'start' in date_range:
                queryset = queryset.filter(**{f"{date_field}__gte": date_range['start']})
            if 'end' in date_range:
                queryset = queryset.filter(**{f"{date_field}__lte": date_range['end']})

        return queryset

    def _apply_group_by(self, queryset, group_by: List[str], aggregations: Dict[str, Any]):
        """Apply group by operations."""
        if aggregations:
            return queryset.values(*group_by).annotate(**aggregations)
        else:
            return queryset.values(*group_by)

    # Export methods for different formats
    def _export_trial_balance_excel(self, data: Dict[str, Any], as_of_date: Optional[date]) -> bytes:
        """Export trial balance to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Trial Balance'

        # Add header
        sheet['A1'] = f'Trial Balance as of {as_of_date or timezone.localdate()}'
        sheet['A3'] = 'Account Code'
        sheet['B3'] = 'Account Name'
        sheet['C3'] = 'Debit Balance'
        sheet['D3'] = 'Credit Balance'

        row = 4
        for category, category_data in data.items():
            if category in ['assets', 'liabilities', 'equity', 'income', 'expenses']:
                sheet[f'A{row}'] = category.title()
                row += 1

                for account in category_data.get('accounts', []):
                    sheet[f'A{row}'] = account['code']
                    sheet[f'B{row}'] = account['name']
                    if category in ['assets', 'expenses']:
                        sheet[f'C{row}'] = float(account['balance'])
                    else:
                        sheet[f'D{row}'] = float(abs(account['balance']))
                    row += 1

                # Total row
                sheet[f'A{row}'] = f'Total {category.title()}'
                sheet[f'C{row}'] = float(category_data['total'])
                row += 2

        # Summary
        sheet[f'A{row}'] = 'Summary'
        sheet[f'A{row+1}'] = 'Total Debits'
        sheet[f'B{row+1}'] = float(data['summary']['total_debit'])
        sheet[f'A{row+2}'] = 'Total Credits'
        sheet[f'B{row+2}'] = float(data['summary']['total_credit'])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_trial_balance_html(self, data: Dict[str, Any], as_of_date: Optional[date]) -> str:
        """Render trial balance as HTML."""
        context = {
            'data': data,
            'as_of_date': as_of_date or timezone.localdate(),
            'organization': self.organization,
        }
        return render_to_string('reports/trial_balance.html', context)

    # Similar methods for other reports...
    def _export_balance_sheet_excel(self, data: Dict[str, Any], as_of_date: Optional[date]) -> bytes:
        """Export balance sheet to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Balance Sheet'

        # Implementation would be similar to trial balance
        # Adding basic structure for now
        sheet['A1'] = f'Balance Sheet as of {as_of_date or timezone.localdate()}'

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_balance_sheet_html(self, data: Dict[str, Any], as_of_date: Optional[date]) -> str:
        """Render balance sheet as HTML."""
        context = {
            'data': data,
            'as_of_date': as_of_date or timezone.localdate(),
            'organization': self.organization,
        }
        return render_to_string('reports/balance_sheet.html', context)

    def _export_income_statement_excel(self, data: Dict[str, Any]) -> bytes:
        """Export income statement to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Income Statement'

        # Implementation would be similar to other reports
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_income_statement_html(self, data: Dict[str, Any]) -> str:
        """Render income statement as HTML."""
        context = {
            'data': data,
            'organization': self.organization,
        }
        return render_to_string('reports/income_statement.html', context)

    def _export_cash_flow_excel(self, data: Dict[str, Any]) -> bytes:
        """Export cash flow to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Cash Flow Statement'

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_cash_flow_html(self, data: Dict[str, Any]) -> str:
        """Render cash flow as HTML."""
        context = {
            'data': data,
            'organization': self.organization,
        }
        return render_to_string('reports/cash_flow.html', context)

    def _export_ledger_excel(self, data: Dict[str, Any]) -> bytes:
        """Export account ledger to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Account Ledger - {data['account']['code']}"

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_ledger_html(self, data: Dict[str, Any]) -> str:
        """Render account ledger as HTML."""
        context = {
            'data': data,
            'organization': self.organization,
        }
        return render_to_string('reports/account_ledger.html', context)

    def _export_tax_report_excel(self, data: Dict[str, Any]) -> bytes:
        """Export tax report to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Tax Report'

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _render_tax_report_html(self, data: Dict[str, Any]) -> str:
        """Render tax report as HTML."""
        context = {
            'data': data,
            'organization': self.organization,
        }
        return render_to_string('reports/tax_report.html', context)

    def _export_custom_report_excel(self, data: Dict[str, Any]) -> bytes:
        """Export custom report to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Custom Report'

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class ReportScheduler:
    """
    Handles scheduled report generation and delivery.
    """

    def __init__(self, organization: Any):
        self.organization = organization

    def schedule_monthly_reports(self, email_recipients: List[str]) -> None:
        """
        Schedule monthly financial reports.

        Args:
            email_recipients: List of email addresses to receive reports
        """
        # This would integrate with Django's task scheduler (Celery, etc.)
        # For now, just define the structure
        monthly_reports = [
            'trial_balance',
            'income_statement',
            'balance_sheet',
            'cash_flow_statement',
        ]

        for report_type in monthly_reports:
            self._schedule_report(
                report_type=report_type,
                frequency='monthly',
                recipients=email_recipients,
                params={'format_type': 'excel'}
            )

    def _schedule_report(self, report_type: str, frequency: str,
                        recipients: List[str], params: Dict[str, Any]) -> None:
        """Schedule individual report."""
        # Implementation would create scheduled tasks
        pass


class ReportTemplateManager:
    """
    Manages custom report templates and configurations.
    """

    def __init__(self, organization: Any):
        self.organization = organization

    def save_report_template(self, name: str, config: Dict[str, Any]) -> None:
        """
        Save a custom report template.

        Args:
            name: Template name
            config: Report configuration
        """
        # This would save to database
        pass

    def get_report_template(self, name: str) -> Optional[Dict[str, Any]]:
        """
        Retrieve a saved report template.

        Args:
            name: Template name

        Returns:
            Template configuration or None
        """
        # This would load from database
        return None

    def list_report_templates(self) -> List[Dict[str, Any]]:
        """
        List all saved report templates.

        Returns:
            List of template metadata
        """
        return []


# Utility functions for common reporting operations
def generate_financial_summary(organization: Any, as_of_date: Optional[date] = None) -> Dict[str, Any]:
    """
    Generate comprehensive financial summary.

    Args:
        organization: Organization instance
        as_of_date: Date for summary

    Returns:
        Financial summary data
    """
    generator = ReportGenerator(organization)

    summary = {
        'trial_balance': generator.generate_trial_balance(as_of_date, format_type='data'),
        'balance_sheet': generator.generate_balance_sheet(as_of_date, format_type='data'),
        'financial_ratios': generator.generate_financial_ratios(as_of_date),
        'generated_at': timezone.now(),
    }

    return summary


def export_report_to_file(report_data: Dict[str, Any], format_type: str, filename: str) -> bytes:
    """
    Export report data to file.

    Args:
        report_data: Report data dictionary
        format_type: Export format
        filename: Output filename

    Returns:
        File content as bytes
    """
    if format_type == 'excel':
        # Convert dict to Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Report'

        # Simple implementation - flatten dict to rows
        row = 1
        for key, value in report_data.items():
            sheet[f'A{row}'] = str(key)
            sheet[f'B{row}'] = str(value)
            row += 1

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    elif format_type == 'json':
        import json
        return json.dumps(report_data, indent=2, default=str).encode('utf-8')

    else:
        raise ValueError(f"Unsupported export format: {format_type}")


def get_report_metadata() -> Dict[str, Dict[str, Any]]:
    """
    Get metadata for all available reports.

    Returns:
        Dictionary of report metadata
    """
    return {
        'trial_balance': {
            'name': 'Trial Balance',
            'description': 'Summary of all account balances',
            'parameters': ['as_of_date', 'include_children'],
            'formats': ['data', 'excel', 'html'],
        },
        'balance_sheet': {
            'name': 'Balance Sheet',
            'description': 'Statement of financial position',
            'parameters': ['as_of_date'],
            'formats': ['data', 'excel', 'html'],
        },
        'income_statement': {
            'name': 'Income Statement',
            'description': 'Statement of profit and loss',
            'parameters': ['start_date', 'end_date'],
            'formats': ['data', 'excel', 'html'],
        },
        'cash_flow': {
            'name': 'Cash Flow Statement',
            'description': 'Statement of cash flows',
            'parameters': ['start_date', 'end_date'],
            'formats': ['data', 'excel', 'html'],
        },
        'account_ledger': {
            'name': 'Account Ledger',
            'description': 'Detailed account transaction history',
            'parameters': ['account_id', 'start_date', 'end_date'],
            'formats': ['data', 'excel', 'html'],
        },
    }
