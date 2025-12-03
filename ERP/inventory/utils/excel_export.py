# Inventory/utils/excel_export.py
"""
Excel Export Utilities for Inventory Management

Generates Excel templates and exports data for:
- Product Master Data
- Price Lists
- Inventory Reports
- Bill of Materials (BOM)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
from typing import List, Optional
from io import BytesIO

from ..models import (
    Product, ProductCategory, Warehouse, Location,
    InventoryItem, PriceList, PriceListItem, StockLedger
)
from enterprise.models import BillOfMaterial


class ExcelStyler:
    """Common styling for Excel files"""
    
    @staticmethod
    def get_header_style():
        return {
            'font': Font(bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    @staticmethod
    def apply_header_style(ws, row_num=1):
        """Apply header styling to first row"""
        styles = ExcelStyler.get_header_style()
        for cell in ws[row_num]:
            cell.font = styles['font']
            cell.fill = styles['fill']
            cell.alignment = styles['alignment']
            cell.border = styles['border']
    
    @staticmethod
    def auto_adjust_columns(ws, min_width=10, max_width=50):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


class ProductExporter:
    """Export product data to Excel"""
    
    @staticmethod
    def create_template() -> BytesIO:
        """Create blank product import template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Import"
        
        # Headers
        headers = [
            'Code', 'Name', 'Category Code', 'Description',
            'Sale Price', 'Cost Price', 'Barcode', 'SKU',
            'Unit of Measure', 'Is Inventory Item', 'Reorder Level', 'Reorder Quantity'
        ]
        ws.append(headers)
        
        # Add example row
        ws.append([
            'PROD001', 'Sample Product', 'CAT01', 'Sample description',
            99.99, 50.00, '1234567890', 'SKU-001',
            'EA', 'Yes', 10, 50
        ])
        
        # Apply styling
        ExcelStyler.apply_header_style(ws)
        ExcelStyler.auto_adjust_columns(ws)
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ['Product Import Template - Instructions'],
            [''],
            ['Required Fields:', 'Code, Name, Category Code'],
            ['Optional Fields:', 'All other fields are optional'],
            [''],
            ['Field Descriptions:'],
            ['Code', 'Unique product code (must be unique)'],
            ['Name', 'Product name'],
            ['Category Code', 'Must match existing category code'],
            ['Sale Price', 'Selling price (decimal)'],
            ['Cost Price', 'Cost/purchase price (decimal)'],
            ['Is Inventory Item', 'Yes/No - whether to track inventory'],
            ['Reorder Level', 'Minimum quantity before reorder alert'],
            ['Reorder Quantity', 'Quantity to order when restocking'],
            [''],
            ['Tips:'],
            ['- Delete the example row before importing'],
            ['- Create categories before importing products'],
            ['- Ensure all codes are unique'],
            ['- Use decimal format for prices (e.g., 99.99)']
        ]
        for row in instructions:
            ws_instructions.append(row)
        
        ws_instructions['A1'].font = Font(bold=True, size=14)
        ExcelStyler.auto_adjust_columns(ws_instructions)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_products(organization, category=None) -> BytesIO:
        """Export existing products to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        headers = [
            'Code', 'Name', 'Category Code', 'Category Name', 'Description',
            'Sale Price', 'Cost Price', 'Barcode', 'SKU',
            'Unit of Measure', 'Is Inventory Item', 'Reorder Level', 'Reorder Quantity'
        ]
        ws.append(headers)
        
        # Query products
        products = Product.objects.filter(organization=organization).select_related('category')
        if category:
            products = products.filter(category=category)
        
        # Add data rows
        for product in products:
            ws.append([
                product.code,
                product.name,
                product.category.code if product.category else '',
                product.category.name if product.category else '',
                product.description,
                float(product.sale_price) if product.sale_price else 0,
                float(product.cost_price) if product.cost_price else 0,
                product.barcode,
                product.sku,
                product.unit_of_measure,
                'Yes' if product.is_inventory_item else 'No',
                float(product.reorder_level) if product.reorder_level else '',
                float(product.reorder_quantity) if product.reorder_quantity else ''
            ])
        
        ExcelStyler.apply_header_style(ws)
        ExcelStyler.auto_adjust_columns(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class PriceListExporter:
    """Export price list data to Excel"""
    
    @staticmethod
    def create_template() -> BytesIO:
        """Create blank price list import template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price List Import"
        
        headers = [
            'Price List Code', 'Product Code', 'Price',
            'Valid From', 'Valid To', 'Min Quantity'
        ]
        ws.append(headers)
        
        # Example row
        ws.append([
            'RETAIL-2025', 'PROD001', 99.99,
            '2025-01-01', '2025-12-31', 1
        ])
        
        ExcelStyler.apply_header_style(ws)
        ExcelStyler.auto_adjust_columns(ws)
        
        # Instructions
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ['Price List Import Template - Instructions'],
            [''],
            ['Required Fields:', 'Price List Code, Product Code, Price'],
            [''],
            ['Field Descriptions:'],
            ['Price List Code', 'Must match existing price list'],
            ['Product Code', 'Must match existing product'],
            ['Price', 'Price for this product (decimal)'],
            ['Valid From', 'Start date (YYYY-MM-DD format)'],
            ['Valid To', 'End date (YYYY-MM-DD format)'],
            ['Min Quantity', 'Minimum order quantity for this price'],
            [''],
            ['Tips:'],
            ['- Create price list header before importing items'],
            ['- Use YYYY-MM-DD format for dates'],
            ['- Leave Valid To blank for no end date'],
            ['- Min Quantity defaults to 1 if blank']
        ]
        for row in instructions:
            ws_instructions.append(row)
        
        ws_instructions['A1'].font = Font(bold=True, size=14)
        ExcelStyler.auto_adjust_columns(ws_instructions)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_price_list(price_list) -> BytesIO:
        """Export price list items to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Price List - {price_list.code}"
        
        # Add price list header info
        ws['A1'] = 'Price List Code:'
        ws['B1'] = price_list.code
        ws['A2'] = 'Price List Name:'
        ws['B2'] = price_list.name
        ws['A3'] = 'Currency:'
        ws['B3'] = price_list.currency_code
        ws.append([])  # Blank row
        
        # Headers
        headers = [
            'Product Code', 'Product Name', 'Price',
            'Valid From', 'Valid To', 'Min Quantity'
        ]
        ws.append(headers)
        
        # Data rows
        items = price_list.items.select_related('product').all()
        for item in items:
            ws.append([
                item.product.code,
                item.product.name,
                float(item.price),
                item.valid_from.strftime('%Y-%m-%d') if item.valid_from else '',
                item.valid_to.strftime('%Y-%m-%d') if item.valid_to else '',
                float(item.min_quantity)
            ])
        
        ExcelStyler.apply_header_style(ws, row_num=5)
        ExcelStyler.auto_adjust_columns(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class InventoryExporter:
    """Export inventory data to Excel"""
    
    @staticmethod
    def create_count_template(organization, warehouse=None) -> BytesIO:
        """Create physical count template with current inventory"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Physical Count"
        
        # Headers
        headers = [
            'Product Code', 'Product Name', 'Warehouse Code',
            'Location Code', 'Batch Number', 'Current Qty',
            'Counted Qty', 'Notes'
        ]
        ws.append(headers)
        
        # Get inventory items
        items = InventoryItem.objects.filter(
            organization=organization
        ).select_related('product', 'warehouse', 'location', 'batch')
        
        if warehouse:
            items = items.filter(warehouse=warehouse)
        
        # Add rows with current quantities
        for item in items:
            ws.append([
                item.product.code,
                item.product.name,
                item.warehouse.code,
                item.location.code if item.location else '',
                item.batch.batch_number if item.batch else '',
                float(item.quantity_on_hand),
                '',  # Blank for manual entry
                ''
            ])
        
        ExcelStyler.apply_header_style(ws)
        ExcelStyler.auto_adjust_columns(ws)
        
        # Protect columns except Counted Qty and Notes
        ws.protection.sheet = True
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                if idx not in [6, 7]:  # Counted Qty and Notes columns
                    cell.protection = openpyxl.styles.Protection(locked=True)
                else:
                    cell.protection = openpyxl.styles.Protection(locked=False)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_stock_report(organization, warehouse=None, as_of_date=None) -> BytesIO:
        """Export current stock levels report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Report"
        
        # Report header
        ws['A1'] = 'Stock Level Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'As of: {as_of_date or datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws.append([])
        
        # Headers
        headers = [
            'Product Code', 'Product Name', 'Category', 'Warehouse',
            'Location', 'Batch', 'Qty On Hand', 'Reorder Level',
            'Status', 'Unit Cost', 'Total Value'
        ]
        ws.append(headers)
        
        # Get inventory data
        items = InventoryItem.objects.filter(
            organization=organization
        ).select_related('product', 'product__category', 'warehouse', 'location', 'batch')
        
        if warehouse:
            items = items.filter(warehouse=warehouse)
        
        total_value = Decimal('0.00')
        
        for item in items:
            qty = item.quantity_on_hand
            cost = item.product.cost_price or Decimal('0.00')
            value = qty * cost
            total_value += value
            
            # Determine status
            status = 'OK'
            if item.product.reorder_level:
                if qty <= 0:
                    status = 'OUT OF STOCK'
                elif qty < item.product.reorder_level:
                    status = 'LOW STOCK'
            
            ws.append([
                item.product.code,
                item.product.name,
                item.product.category.name if item.product.category else '',
                item.warehouse.code,
                item.location.code if item.location else '',
                item.batch.batch_number if item.batch else '',
                float(qty),
                float(item.product.reorder_level) if item.product.reorder_level else '',
                status,
                float(cost),
                float(value)
            ])
        
        # Add total row
        last_row = ws.max_row + 1
        ws[f'J{last_row}'] = 'TOTAL:'
        ws[f'J{last_row}'].font = Font(bold=True)
        ws[f'K{last_row}'] = float(total_value)
        ws[f'K{last_row}'].font = Font(bold=True)
        
        ExcelStyler.apply_header_style(ws, row_num=4)
        ExcelStyler.auto_adjust_columns(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class BOMExporter:
    """Export Bill of Materials to Excel"""
    
    @staticmethod
    def create_template() -> BytesIO:
        """Create BOM import template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM Import"
        
        headers = [
            'Parent Code', 'Component Code', 'Quantity',
            'Unit of Measure', 'Scrap %', 'Notes'
        ]
        ws.append(headers)
        
        # Example rows for one BOM
        ws.append(['FG-001', 'RM-001', 2.5, 'KG', 5.0, 'Main ingredient'])
        ws.append(['FG-001', 'RM-002', 1.0, 'L', 2.0, 'Liquid component'])
        ws.append(['FG-001', 'PKG-001', 1.0, 'EA', 0.0, 'Packaging'])
        
        ExcelStyler.apply_header_style(ws)
        ExcelStyler.auto_adjust_columns(ws)
        
        # Instructions
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ['BOM Import Template - Instructions'],
            [''],
            ['Required Fields:', 'Parent Code, Component Code, Quantity'],
            [''],
            ['Field Descriptions:'],
            ['Parent Code', 'Finished good product code'],
            ['Component Code', 'Raw material/component product code'],
            ['Quantity', 'Quantity required per unit of parent'],
            ['Unit of Measure', 'Unit for component quantity'],
            ['Scrap %', 'Expected scrap/waste percentage'],
            ['Notes', 'Additional notes or instructions'],
            [''],
            ['Tips:'],
            ['- Group all components for same parent together'],
            ['- All product codes must exist before import'],
            ['- Scrap % is optional, defaults to 0'],
            ['- Use decimal format for quantities']
        ]
        for row in instructions:
            ws_instructions.append(row)
        
        ws_instructions['A1'].font = Font(bold=True, size=14)
        ExcelStyler.auto_adjust_columns(ws_instructions)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_bom(bom) -> BytesIO:
        """Export BOM to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"BOM - {bom.product.code}"
        
        # BOM header
        ws['A1'] = 'Product Code:'
        ws['B1'] = bom.product.code
        ws['A2'] = 'Product Name:'
        ws['B2'] = bom.product.name
        ws['A3'] = 'BOM Version:'
        ws['B3'] = bom.version
        ws.append([])
        
        # Headers
        headers = [
            'Line', 'Component Code', 'Component Name', 'Quantity',
            'Unit of Measure', 'Scrap %', 'Notes'
        ]
        ws.append(headers)
        
        # BOM lines
        for idx, line in enumerate(bom.lines.select_related('component').all(), start=1):
            ws.append([
                idx,
                line.component.code,
                line.component.name,
                float(line.quantity),
                line.unit_of_measure,
                float(line.scrap_percent) if line.scrap_percent else 0,
                line.notes
            ])
        
        ExcelStyler.apply_header_style(ws, row_num=5)
        ExcelStyler.auto_adjust_columns(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
