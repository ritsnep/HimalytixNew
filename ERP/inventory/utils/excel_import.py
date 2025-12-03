# Inventory/utils/excel_import.py
"""
Excel Import Handlers for Inventory Management

Supports bulk import of:
- Product Master Data
- Price Lists
- Inventory Adjustments
- Bill of Materials (BOM)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.db import transaction
from django.core.exceptions import ValidationError
from typing import List, Dict, Tuple

from ..models import (
    Product, ProductCategory, Warehouse, Location,
    InventoryItem, PriceList, PriceListItem, Batch
)
from enterprise.models import BillOfMaterial, BillOfMaterialItem


class ImportResult:
    """Container for import operation results"""
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
    
    def add_error(self, row: int, field: str, message: str):
        self.error_count += 1
        self.errors.append({'row': row, 'field': field, 'message': message})
    
    def add_warning(self, row: int, message: str):
        self.warnings.append({'row': row, 'message': message})
    
    def add_success(self):
        self.success_count += 1
    
    def to_dict(self):
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'errors': self.errors,
            'warnings': self.warnings
        }


class BaseExcelImporter:
    """Base class for Excel import operations"""
    
    def __init__(self, organization, user_id=None):
        self.organization = organization
        self.user_id = user_id
        self.result = ImportResult()
    
    def parse_decimal(self, value, row: int, field: str) -> Decimal:
        """Parse decimal value with error handling"""
        if value is None or value == '':
            return Decimal('0.00')
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            self.result.add_error(row, field, f"Invalid decimal value: {value}")
            return None
    
    def parse_date(self, value, row: int, field: str):
        """Parse date value with error handling"""
        if value is None or value == '':
            return None
        
        if isinstance(value, (date, datetime)):
            return value.date() if isinstance(value, datetime) else value
        
        try:
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(str(value), fmt).date()
                except ValueError:
                    continue
            raise ValueError(f"Unrecognized date format: {value}")
        except Exception as e:
            self.result.add_error(row, field, str(e))
            return None
    
    def parse_boolean(self, value) -> bool:
        """Parse boolean value"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['yes', 'true', '1', 'y']
        return bool(value)


class ProductImporter(BaseExcelImporter):
    """Import products from Excel file"""
    
    REQUIRED_COLUMNS = ['Code', 'Name', 'Category Code']
    OPTIONAL_COLUMNS = [
        'Description', 'Sale Price', 'Cost Price', 'Barcode', 'SKU',
        'Unit of Measure', 'Is Inventory Item', 'Reorder Level', 'Reorder Quantity'
    ]
    
    def import_from_file(self, file_path: str) -> ImportResult:
        """Import products from Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Validate headers
        headers = [cell.value for cell in ws[1]]
        if not all(col in headers for col in self.REQUIRED_COLUMNS):
            self.result.add_error(1, 'Headers', f"Missing required columns: {self.REQUIRED_COLUMNS}")
            return self.result
        
        # Create header mapping
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        # Process rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[col_map['Code']]:  # Skip empty rows
                continue
            
            try:
                self._import_product_row(row, col_map, row_idx)
            except Exception as e:
                self.result.add_error(row_idx, 'General', str(e))
        
        return self.result
    
    @transaction.atomic
    def _import_product_row(self, row: tuple, col_map: dict, row_idx: int):
        """Import a single product row"""
        code = row[col_map['Code']]
        name = row[col_map['Name']]
        category_code = row[col_map['Category Code']]
        
        # Get or validate category
        try:
            category = ProductCategory.objects.get(
                organization=self.organization,
                code=category_code
            )
        except ProductCategory.DoesNotExist:
            self.result.add_error(row_idx, 'Category Code', f"Category not found: {category_code}")
            return
        
        # Parse optional fields
        sale_price = self.parse_decimal(
            row[col_map.get('Sale Price', len(row))], row_idx, 'Sale Price'
        ) or Decimal('0.00')
        
        cost_price = self.parse_decimal(
            row[col_map.get('Cost Price', len(row))], row_idx, 'Cost Price'
        ) or Decimal('0.00')
        
        reorder_level = self.parse_decimal(
            row[col_map.get('Reorder Level', len(row))], row_idx, 'Reorder Level'
        )
        
        reorder_qty = self.parse_decimal(
            row[col_map.get('Reorder Quantity', len(row))], row_idx, 'Reorder Quantity'
        )
        
        is_inventory = self.parse_boolean(
            row[col_map.get('Is Inventory Item', len(row))] if 'Is Inventory Item' in col_map else True
        )
        
        # Create or update product
        product, created = Product.objects.update_or_create(
            organization=self.organization,
            code=code,
            defaults={
                'name': name,
                'category': category,
                'description': row[col_map.get('Description', len(row))] or '',
                'sale_price': sale_price,
                'cost_price': cost_price,
                'barcode': row[col_map.get('Barcode', len(row))] or '',
                'sku': row[col_map.get('SKU', len(row))] or '',
                'unit_of_measure': row[col_map.get('Unit of Measure', len(row))] or 'EA',
                'is_inventory_item': is_inventory,
                'reorder_level': reorder_level,
                'reorder_quantity': reorder_qty,
            }
        )
        
        self.result.add_success()
        if not created:
            self.result.add_warning(row_idx, f"Product {code} updated (already existed)")


class PriceListImporter(BaseExcelImporter):
    """Import price list items from Excel file"""
    
    REQUIRED_COLUMNS = ['Price List Code', 'Product Code', 'Price']
    OPTIONAL_COLUMNS = ['Valid From', 'Valid To', 'Min Quantity']
    
    def import_from_file(self, file_path: str) -> ImportResult:
        """Import price list items from Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Validate headers
        headers = [cell.value for cell in ws[1]]
        if not all(col in headers for col in self.REQUIRED_COLUMNS):
            self.result.add_error(1, 'Headers', f"Missing required columns: {self.REQUIRED_COLUMNS}")
            return self.result
        
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        # Process rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[col_map['Price List Code']]:
                continue
            
            try:
                self._import_price_row(row, col_map, row_idx)
            except Exception as e:
                self.result.add_error(row_idx, 'General', str(e))
        
        return self.result
    
    @transaction.atomic
    def _import_price_row(self, row: tuple, col_map: dict, row_idx: int):
        """Import a single price list item row"""
        price_list_code = row[col_map['Price List Code']]
        product_code = row[col_map['Product Code']]
        price = self.parse_decimal(row[col_map['Price']], row_idx, 'Price')
        
        if price is None:
            return
        
        # Get price list
        try:
            price_list = PriceList.objects.get(
                organization=self.organization,
                code=price_list_code
            )
        except PriceList.DoesNotExist:
            self.result.add_error(row_idx, 'Price List Code', f"Price list not found: {price_list_code}")
            return
        
        # Get product
        try:
            product = Product.objects.get(
                organization=self.organization,
                code=product_code
            )
        except Product.DoesNotExist:
            self.result.add_error(row_idx, 'Product Code', f"Product not found: {product_code}")
            return
        
        # Parse optional fields
        valid_from = self.parse_date(
            row[col_map.get('Valid From', len(row))], row_idx, 'Valid From'
        )
        valid_to = self.parse_date(
            row[col_map.get('Valid To', len(row))], row_idx, 'Valid To'
        )
        min_qty = self.parse_decimal(
            row[col_map.get('Min Quantity', len(row))], row_idx, 'Min Quantity'
        ) or Decimal('1.00')
        
        # Create or update price list item
        item, created = PriceListItem.objects.update_or_create(
            price_list=price_list,
            product=product,
            defaults={
                'price': price,
                'valid_from': valid_from,
                'valid_to': valid_to,
                'min_quantity': min_qty
            }
        )
        
        self.result.add_success()


class InventoryAdjustmentImporter(BaseExcelImporter):
    """Import inventory adjustments from Excel file"""
    
    REQUIRED_COLUMNS = ['Product Code', 'Warehouse Code', 'Quantity']
    OPTIONAL_COLUMNS = ['Location Code', 'Batch Number', 'Notes']
    
    def import_from_file(self, file_path: str, adjustment_type: str = 'count') -> ImportResult:
        """Import inventory adjustments from Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if not all(col in headers for col in self.REQUIRED_COLUMNS):
            self.result.add_error(1, 'Headers', f"Missing required columns: {self.REQUIRED_COLUMNS}")
            return self.result
        
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[col_map['Product Code']]:
                continue
            
            try:
                self._import_adjustment_row(row, col_map, row_idx, adjustment_type)
            except Exception as e:
                self.result.add_error(row_idx, 'General', str(e))
        
        return self.result
    
    @transaction.atomic
    def _import_adjustment_row(self, row: tuple, col_map: dict, row_idx: int, adjustment_type: str):
        """Import a single inventory adjustment row"""
        product_code = row[col_map['Product Code']]
        warehouse_code = row[col_map['Warehouse Code']]
        quantity = self.parse_decimal(row[col_map['Quantity']], row_idx, 'Quantity')
        
        if quantity is None:
            return
        
        # Get product
        try:
            product = Product.objects.get(
                organization=self.organization,
                code=product_code
            )
        except Product.DoesNotExist:
            self.result.add_error(row_idx, 'Product Code', f"Product not found: {product_code}")
            return
        
        # Get warehouse
        try:
            warehouse = Warehouse.objects.get(
                organization=self.organization,
                code=warehouse_code
            )
        except Warehouse.DoesNotExist:
            self.result.add_error(row_idx, 'Warehouse Code', f"Warehouse not found: {warehouse_code}")
            return
        
        # Get optional location
        location = None
        location_code = row[col_map.get('Location Code', len(row))]
        if location_code:
            try:
                location = Location.objects.get(
                    warehouse=warehouse,
                    code=location_code
                )
            except Location.DoesNotExist:
                self.result.add_warning(row_idx, f"Location not found: {location_code}, using warehouse default")
        
        # Get optional batch
        batch = None
        batch_number = row[col_map.get('Batch Number', len(row))]
        if batch_number:
            batch, _ = Batch.objects.get_or_create(
                organization=self.organization,
                product=product,
                batch_number=batch_number
            )
        
        # Create or update inventory item
        inventory_item, created = InventoryItem.objects.get_or_create(
            organization=self.organization,
            product=product,
            warehouse=warehouse,
            location=location,
            batch=batch,
            defaults={'quantity_on_hand': Decimal('0.00')}
        )
        
        if adjustment_type == 'count':
            # Physical count - set to exact quantity
            inventory_item.quantity_on_hand = quantity
        else:
            # Adjustment - add/subtract
            inventory_item.quantity_on_hand += quantity
        
        inventory_item.save()
        
        self.result.add_success()


class BOMImporter(BaseExcelImporter):
    """Import Bill of Materials from Excel file"""
    
    REQUIRED_COLUMNS = ['Parent Code', 'Component Code', 'Quantity']
    OPTIONAL_COLUMNS = ['Unit of Measure', 'Scrap %', 'Notes']
    
    def import_from_file(self, file_path: str) -> ImportResult:
        """Import BOM from Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if not all(col in headers for col in self.REQUIRED_COLUMNS):
            self.result.add_error(1, 'Headers', f"Missing required columns: {self.REQUIRED_COLUMNS}")
            return self.result
        
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        # Group rows by parent
        bom_groups = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            parent_code = row[col_map['Parent Code']]
            if not parent_code:
                continue
            
            if parent_code not in bom_groups:
                bom_groups[parent_code] = []
            bom_groups[parent_code].append((row, row_idx))
        
        # Process each BOM
        for parent_code, rows in bom_groups.items():
            try:
                self._import_bom(parent_code, rows, col_map)
            except Exception as e:
                self.result.add_error(rows[0][1], 'General', str(e))
        
        return self.result
    
    @transaction.atomic
    def _import_bom(self, parent_code: str, rows: List[Tuple], col_map: dict):
        """Import BOM for a single parent product"""
        # Get parent product
        try:
            parent_product = Product.objects.get(
                organization=self.organization,
                code=parent_code
            )
        except Product.DoesNotExist:
            self.result.add_error(rows[0][1], 'Parent Code', f"Product not found: {parent_code}")
            return
        
        # Create BOM
        bom, created = BillOfMaterial.objects.get_or_create(
            organization=self.organization,
            product=parent_product,
            defaults={
                'bom_number': f"BOM-{parent_code}",
                'version': 1,
                'is_active': True
            }
        )
        
        # Delete existing lines if updating
        if not created:
            bom.lines.all().delete()
        
        # Create BOM lines
        for row, row_idx in rows:
            component_code = row[col_map['Component Code']]
            quantity = self.parse_decimal(row[col_map['Quantity']], row_idx, 'Quantity')
            
            if quantity is None:
                continue
            
            # Get component product
            try:
                component = Product.objects.get(
                    organization=self.organization,
                    code=component_code
                )
            except Product.DoesNotExist:
                self.result.add_error(row_idx, 'Component Code', f"Product not found: {component_code}")
                continue
            
            scrap_percent = self.parse_decimal(
                row[col_map.get('Scrap %', len(row))], row_idx, 'Scrap %'
            ) or Decimal('0.00')
            
            BillOfMaterialItem.objects.create(
                bom=bom,
                component=component,
                quantity=quantity,
                unit_of_measure=row[col_map.get('Unit of Measure', len(row))] or 'EA',
                scrap_percent=scrap_percent,
                notes=row[col_map.get('Notes', len(row))] or ''
            )
            
            self.result.add_success()
